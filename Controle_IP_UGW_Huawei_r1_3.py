# coding=utf-8
import paramiko
import os
from funcoes import find_before, find_between, find_after
import datetime, time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# -----------------------------------------------------------------------------
# inicia uma sessão SSH
def ssh_session_UGW(UGW_name, UGW_ip):
    print "Conectando ao " + UGW_name

    if UGW_name == 'GPRJO3' or UGW_name == 'GPRJO4' or UGW_name == 'GPSPO3' or UGW_name == 'GPSNE1':
        UGW_senha = 'CNEPS@tim02'
    else:
        UGW_senha = 'CNEPS@tim01'

    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        ssh.connect(UGW_ip, username='F8019834', password=str(UGW_senha))
    except paramiko.SSHException:
        print "Connection Failed"
        quit()

    resposta = "[" + UGW_name + "] "
    resp = ""

    chan = ssh.invoke_shell()
    # time.sleep(3)
    while resp <> ">":
        resp = chan.recv(1)
    resp = ""

    # Send command
    chan.send('system-view\n')
    while resposta[-8:] <> "[" + UGW_name + "]":
        resp = chan.recv(1)
        resposta = resposta + resp
    resposta = "[" + UGW_name + "] "
    resp = ""

    # Send command
    print "     printing: display current-configuration"
    chan.send('display current-configuration\n')
    while resposta[-8:] <> "[" + UGW_name + "]":
        resp = chan.recv(1)
        resposta = resposta + resp
        if resposta[-9:-5] == "More":
            chan.send(' ')
    f = open('./Prints_UGW/' + UGW_name + '_print.txt', 'wb')
    f.write(resposta)
    f.close()

    resposta = "[" + UGW_name + "] "
    resp = ""

    # Send command
    chan.send('quit\n')
    while resposta[-8:] <> "<" + UGW_name + ">":
        resp = chan.recv(1)
        resposta = resposta + resp

    resposta = "[" + UGW_name + "] "
    resp = ""

    # remover os ---- MORE ---- dos prints
    f = open('./Prints_UGW/' + UGW_name + '_print.txt')
    lines = f.readlines()
    f.close()
    novo_arq = ""
    for line in lines:
        if line[0:68] == '  ---- More ----[42D                                          [42D':
            line = line.replace('  ---- More ----[42D                                          [42D', '')
        novo_arq += line
    f = open('./Prints_UGW/' + UGW_name + '_print.txt', 'wb')
    f.write(novo_arq)
    f.close()

    print "     Concluido."
    print

    ssh.close()
    return ()


# -----------------------------------------------------------------

# -----------------------------------------------------------------
def pode14(octeto):
    while octeto > 0: octeto -= 4
    if octeto == 0:
        return True
    else:
        return False


# -----------------------------------------------------------------
# -----------------------------------------------------------------
def pode15(octeto):
    while octeto > 0: octeto -= 2
    if octeto == 0:
        return True
    else:
        return False


# -----------------------------------------------------------------
# -----------------------------------------------------------------
def pode17(octeto):
    if octeto == 0:
        return True
    else:
        if octeto == 128:
            return True
        else:
            return False


# -----------------------------------------------------------------
# -----------------------------------------------------------------
def pode18(octeto):
    while octeto > 0: octeto -= 64
    if octeto == 0:
        return True
    else:
        return False


# -----------------------------------------------------------------
# -----------------------------------------------------------------
def pode19(octeto):
    while octeto > 0: octeto -= 32
    if octeto == 0:
        return True
    else:
        return False


# -----------------------------------------------------------------
# -----------------------------------------------------------------
def pode20(octeto):
    while octeto > 0: octeto -= 16
    if octeto == 0:
        return True
    else:
        return False


# -----------------------------------------------------------------
# -----------------------------------------------------------------
def pode21(octeto):
    while octeto > 0: octeto -= 8
    if octeto == 0:
        return True
    else:
        return False


# -----------------------------------------------------------------
# -----------------------------------------------------------------
def pode22(octeto):
    while octeto > 0: octeto -= 4
    if octeto == 0:
        return True
    else:
        return False


# -----------------------------------------------------------------

# -----------------------------------------------------------------
def sumarizar(elemento):
    f = open('./Prints_UGW/ip-pool/' + elemento + '_ip-range-list.txt')
    lines = f.readlines()
    f.close()
    # print
    # print 'elemento: ' + elemento
    linhas = len(lines)
    # print '\nLinhas: ' + str(linhas)
    lista_rede = []
    linha = 0
    while linha < linhas - 1:
        line_A = lines[linha]
        line_B = lines[linha + 1]
        # print '\n... linha A: ' + line_A + '--- linha B: ' + line_B
        start = 0
        end = line_A.index(".", start)
        octeto1A = line_A[start:end]
        # print '... octeto A1: ' + octeto1A
        start = end + 1
        end = line_A.index(".", start)
        octeto2A = line_A[start:end]
        # print '... octeto A2: ' + octeto2A
        start = end + 1
        end = line_A.index(".", start)
        octeto3A = line_A[start:end]
        # print '... octeto A3: ' + octeto3A
        start = end + 1
        end = line_A.index("/", start)
        octeto4A = line_A[start:end]
        # print '... octeto A4: ' + octeto4A
        start = end + 1
        mask_A = line_A[start:-1]
        # print '... MASK A: ' + mask_A

        start = 0
        end = line_B.index(".", start)
        octeto1B = line_B[start:end]
        # print '--- octeto B1: ' + octeto1B
        start = end + 1
        end = line_B.index(".", start)
        octeto2B = line_B[start:end]
        # print '--- octeto B2: ' + octeto2B
        start = end + 1
        end = line_B.index(".", start)
        octeto3B = line_B[start:end]
        # print '--- octeto B3: ' + octeto3B
        start = end + 1
        end = line_B.index("/", start)
        octeto4B = line_B[start:end]
        # print '--- octeto B4: ' + octeto4B
        start = end + 1
        mask_B = line_B[start:-1]
        # print '--- MASK B: ' + mask_B
        linha += 1

        # verificar se rede_A pode simarizar com rede_B
        if len(mask_A) > 2 or len(mask_B) > 2:
            lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
            continue
        if octeto1A != octeto1B:
            lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
            continue
        if int(mask_A) < 15:
            lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
            continue
        if int(mask_A) != int(mask_B):
            lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
            continue
        if int(mask_A) == 15:
            if pode14(int(octeto2A)) != True:
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
            if int(octeto2A) != int(octeto2B) - 2:
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
            else:
                lista_rede.append(octeto1A + '.' + octeto2A + '.0.0/' + str(int(mask_A) - 1) + '\n')
                # print "pode sumarizar: "+octeto1A+'.'+octeto2A+'.'+octeto3A+'.'+octeto4A+'/'+str(int(mask_A)-1)
                linha += 1
                continue
        if int(mask_A) == 16:
            if pode15(int(octeto2A)) != True:
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
            if int(octeto2A) != int(octeto2B) - 1:
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
            else:
                lista_rede.append(octeto1A + '.' + octeto2A + '.0.0/' + str(int(mask_A) - 1) + '\n')
                # print "pode sumarizar: "+octeto1A+'.'+octeto2A+'.'+octeto3A+'.'+octeto4A+'/'+str(int(mask_A)-1)
                linha += 1
                continue
        if int(mask_A) == 17:
            if int(octeto2A) != int(octeto2B):
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
            else:
                lista_rede.append(octeto1A + '.' + octeto2A + '.0.0/' + str(int(mask_A) - 1) + '\n')
                # print "pode sumarizar: "+octeto1A+'.'+octeto2A+'.'+octeto3A+'.'+octeto4A+'/'+str(int(mask_A)-1)
                linha += 1
                continue
        if int(mask_A) == 18:
            if pode17(int(octeto3A)) != True:
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
            if int(octeto2A) != int(octeto2B):
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
            if int(octeto2A) == int(octeto2B) and int(octeto3A) == int(octeto3B) - 64:
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.0/' + str(int(mask_A) - 1) + '\n')
                # print "pode sumarizar: "+octeto1A+'.'+octeto2A+'.'+octeto3A+'.'+octeto4A+'/'+str(int(mask_A)-1)
                linha += 1
                continue
            else:
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
        if int(mask_A) == 19:
            if pode18(int(octeto3A)) != True:
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
            if int(octeto2A) != int(octeto2B):
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
            if int(octeto2A) == int(octeto2B) and int(octeto3A) == int(octeto3B) - 32:
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.0/' + str(int(mask_A) - 1) + '\n')
                # print "pode sumarizar: "+octeto1A+'.'+octeto2A+'.'+octeto3A+'.'+octeto4A+'/'+str(int(mask_A)-1)
                linha += 1
                continue
            else:
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
        if int(mask_A) == 20:
            if pode19(int(octeto3A)) != True:
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
            if int(octeto2A) != int(octeto2B):
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
            if int(octeto2A) == int(octeto2B) and int(octeto3A) == int(octeto3B) - 16:
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.0/' + str(int(mask_A) - 1) + '\n')
                # print "pode sumarizar: "+octeto1A+'.'+octeto2A+'.'+octeto3A+'.'+octeto4A+'/'+str(int(mask_A)-1)
                linha += 1
                continue
            else:
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
        if int(mask_A) == 21:
            if pode20(int(octeto3A)) != True:
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
            if int(octeto2A) != int(octeto2B):
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
            if int(octeto2A) == int(octeto2B) and int(octeto3A) == int(octeto3B) - 8:
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.0/' + str(int(mask_A) - 1) + '\n')
                # print "pode sumarizar: "+octeto1A+'.'+octeto2A+'.'+octeto3A+'.'+octeto4A+'/'+str(int(mask_A)-1)
                linha += 1
                continue
            else:
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
        if int(mask_A) == 22:
            if pode21(int(octeto3A)) != True:
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
            if int(octeto2A) != int(octeto2B):
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
            if int(octeto2A) == int(octeto2B) and int(octeto3A) == int(octeto3B) - 4:
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.0/' + str(int(mask_A) - 1) + '\n')
                # print "pode sumarizar: "+octeto1A+'.'+octeto2A+'.'+octeto3A+'.'+octeto4A+'/'+str(int(mask_A)-1)
                linha += 1
                continue
            else:
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
        if int(mask_A) == 23:
            if pode22(int(octeto3A)) != True:
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
            if int(octeto2A) != int(octeto2B):
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
            if int(ocline_Ateto2A) == int(octeto2B) and int(octeto3A) == int(octeto3B) - 2:
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.0/' + str(int(mask_A) - 1) + '\n')
                # print "pode sumarizar: "+octeto1A+'.'+octeto2A+'.'+octeto3A+'.'+octeto4A+'/'+str(int(mask_A)-1)
                linha += 1
                continue
            else:
                lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
                continue
        if int(mask_A) == 24 and octeto2A == octeto2B and int(octeto3A) == int(octeto3B) - 1:
            lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.0/' + str(int(mask_A) - 1) + '\n')
            # print "pode sumarizar: "+octeto1A+'.'+octeto2A+'.'+octeto3A+'.'+octeto4A+'/'+str(int(mask_A)-1)
            linha += 1
            continue
        else:
            lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')
            continue

    if linha < linhas:
        line_A = lines[linha]
        start = 0
        end = line_A.index(".", start)
        octeto1A = line_A[start:end]
        start = end + 1
        end = line_A.index(".", start)
        octeto2A = line_A[start:end]
        start = end + 1
        end = line_A.index(".", start)
        octeto3A = line_A[start:end]
        start = end + 1
        end = line_A.index("/", start)
        octeto4A = line_A[start:end]
        start = end + 1
        mask_A = line_A[start:-1]

        lista_rede.append(octeto1A + '.' + octeto2A + '.' + octeto3A + '.' + octeto4A + '/' + mask_A + '\n')

    f = open('./Prints_UGW/ip-pool/' + elemento + '_ip-range-list.txt', 'wb')
    for line in lista_rede:
        f.write(line)
    f.close()

    return ()


# -----------------------------------------------------------------
# -----------------------------------------------------------------
def find_rede_huawei(texto):
    end = texto.index(".", 0)
    octeto1A = texto[:end]
    # print 'A1='+octeto1A

    start = end + 1
    end = texto.index(".", start)
    octeto2A = texto[start:end]
    # print 'A1='+octeto2A

    start = end + 1
    end = texto.index(".", start)
    octeto3A = texto[start:end]
    # print 'A1='+octeto3A

    start = end + 1
    end = texto.index("-", start)
    octeto4A = texto[start:end]

    start = end + 1
    end = texto.index(".", start)

    start = end + 1
    end = texto.index(".", start)

    start = end + 1
    end = texto.index(".", start)
    octeto3B = texto[start:end]
    # print 'B3='+octeto3B

    tamanho_range = (int(octeto3B) - int(octeto3A) + 1) * 256

    mask = 25
    while tamanho_range > (2 ** (32 - mask)):
        mask = mask - 1

    if int(octeto1A) < 100: octeto1A = "0" + octeto1A
    if int(octeto1A) < 10: octeto1A = "0" + octeto1A
    if int(octeto2A) < 100: octeto2A = "0" + octeto2A
    if int(octeto2A) < 10: octeto2A = "0" + octeto2A
    if int(octeto3A) < 100: octeto3A = "0" + octeto3A
    if int(octeto3A) < 10: octeto3A = "0" + octeto3A

    rede = octeto1A + '.' + octeto2A + '.' + octeto3A + '.0/' + str(mask) + '\n'
    # print rede
    return rede


# -----------------------------------------------------------------

# -----------------------------------------------------------------
def salva_todos(elemento):
    f = open('./Prints_UGW/ip-pool/' + elemento + '_ip-range-list.txt')
    lista = f.readlines()
    f.close()

    f = open('./Prints_UGW/ip-pool/todos_rede.txt', 'ab')
    for line in lista:
        f.write(line[:-1] + '-' + elemento + '\n')
        # print line
    f.close()


# -----------------------------------------------------------------
# -----------------------------------------------------------------
def ordena_todos():
    f = open('./Prints_UGW/ip-pool/todos_rede.txt')
    lista = f.readlines()
    f.close()
    lista.sort()
    f = open('./Prints_UGW/ip-pool/todos_rede.txt', 'wb')
    for line in lista:
        end = line.index(".", 0)
        octeto1A = str(int(line[:end]))
        start = end + 1
        end = line.index(".", start)
        octeto2A = str(int(line[start:end]))
        start = end + 1
        end = line.index(".", start)
        octeto3A = str(int(line[start:end]))
        start = end + 1
        end = line.index("/", start)
        octeto4A = str(int(line[start:end]))
        start = end + 1
        mask = line[start:-1]
        line = octeto1A + "." + octeto2A + "." + octeto3A + ".0/" + str(mask) + '\n'
        f.write(line)
    f.close()


# -----------------------------------------------------------------




#                      #
##                    ##
### Função principal ###
##                    ##
#                      #



###### Definir quais bases de FNGs vão ser atualizados


lista_UGWs = {'GPBHE1': '10.207.102.244',
              'GPBLM1': '10.207.99.244',
              'GPBSA1': '10.207.100.244',
              'GPCTA1': '10.207.101.244',
              'GPMNS1': '10.207.104.244',
              'GPRCE1': '10.207.98.244',
              'GPRJO3': '10.207.105.244',
              'GPRJO4': '10.207.106.244',
              'GPSDR1': '10.207.103.244',
              'GPSNE1': '10.207.107.244',
              'GPSPO3': '10.207.108.244',
              'BBIP': '1.1.1.1'}

# --------------------------------
# Coleta dos prints nos UGWs
print
print("Coletando dados dos UGWs:")
print
if not os.path.exists('Prints_UGW'): os.mkdir('Prints_UGW')

UGWs = raw_input("UGWs (ex.: GPBSA1;GPBHE1..., <vazio> para todos ou <no> para não atualizar): ")
if UGWs == "":
    UGWs = "GPBHE1;GPBLM1;GPBSA1;GPCTA1;GPMNS1;GPRCE1;GPRJO3;GPRJO4;GPSDR1;GPSNE1;GPSPO3"
else:
    if UGWs.upper() == "NO":
        UGWs = ""
UGWs = UGWs.split(';')

if UGWs[0] != "":
    try:
        for UGW in UGWs:
            ssh_session_UGW(UGW, lista_UGWs[UGW])
    except:
        print(u'Falhou conexão com ' + UGW)
        print
# --------------------------------














# --------------------------------
# identificar os IP-POOL

### Zera o arquivo "./Prints_UGW/ip-pool/todos_rede.txt"
f = open('./Prints_UGW/ip-pool/todos_rede.txt', 'wb')
f.close()

# for UGW in ['GPCTA1','GPRJO3']:
for UGW in lista_UGWs:
    f = open('./Prints_UGW/' + UGW + '_print.txt')
    lines = f.readlines()
    f.close()
    iprange = []
    iprangelist = []
    ippoollist = ""
    lenlines = len(lines)
    for line in range(lenlines - 1):
        if ('ip pool timbrasil' in lines[line][:30] and 'ip pool timbrasil-ipv6' not in lines[line][:40]) or (
            'ip pool timip' in lines[line][:30]) or (
                'ip pool pool-wvpn' in lines[line][:30] and 'ip pool pool-wvpnblock' not in lines[line][:40]):
            line += 1
            while 'alarm-report' not in lines[line]:
                if ('section' in lines[line] and 'lock' not in lines[line]) and ('lock' not in lines[line + 1]):
                    # print
                    # print ['UGW', UGW]
                    # print ['line',line,lines[line]]
                    iprange = lines[line].split(" ")
                    # print ['iprange',iprange]
                    iprangelist.append(find_rede_huawei(iprange[4] + "-" + iprange[5]))
                    # print ['iprangelist',iprangelist]
                    ippoollist += lines[line]
                    # print ['ippoollist',ippoollist]
                    # print

                if line < lenlines - 2:
                    line += 1
                else:
                    break
    f = open('./Prints_UGW/ip-pool/' + UGW + '_ip-pool-list.txt', 'wb')
    f.write(ippoollist)
    f.close()
    s = ''
    iprangelist.sort()
    for a in range(len(iprangelist)):
        s += iprangelist[a]
    f = open('./Prints_UGW/ip-pool/' + UGW + '_ip-range-list.txt', 'wb')
    f.write(s)
    f.close()
    for a in range(0, 10): sumarizar(UGW)
    salva_todos(UGW)
ordena_todos()

# quit()





# montar relatório
f = open('./Prints_UGW/ip-pool/todos_rede.txt')
lista_A = f.readlines()
f.close()

f = open('./Prints_UGW/ip-pool/Lista_IPs_publicos_sub21.txt')
lista_B = f.readlines()
f.close()
for i in range(0, len(lista_A)):
    rede_A = find_before(lista_A[i], '/')
    mask_A = find_between(lista_A[i], '/', '-')
    elemento = find_after(lista_A[i], '-')
    j = 0
    while j < len(lista_B):
        rede_B = find_before(lista_B[j], '/')
        if rede_B == rede_A:
            break
        else:
            j += 1
    if j == len(lista_B):
        print "\n\n\nconflito..."
        print 'rede: ' + rede_A
        print 'elemento: ' + elemento
        # time.sleep(6)
        print
        confirma = raw_input("Tecle algo para terminar...")
        print "Concluido..."
        quit()
    s = {'20': 2, '19': 4, '18': 8, '17': 16, '16': 32, '15': 64, '14': 128}
    if mask_A == '21':
        lista_B[j] = lista_B[j][:-1] + '-' + elemento
    else:
        lista_B[j] = rede_A + '/' + mask_A + '-' + elemento
        for k in range(j + (s[mask_A]) - 1, j, -1):
            lista_B.remove(lista_B[k])
f = open('./Prints_UGW/ip-pool/IPs_publicos_ip-range-list.txt', 'wb')
for line in lista_B:
    f.write(line)
f.close()
for a in range(0, 10): sumarizar('IPs_publicos')
# Criar planilha resultado
wb = Workbook()
ws = wb.create_sheet(0, title='Ranges')
ws.sheet_properties.tabColor = "107272"
sheet = ws
sheet['A1'] = 'Range IP'
sheet['A1'].font = Font(name='Courrier', size=11, bold=False)
sheet['A1'].alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False,
                                  shrink_to_fit=False, indent=0)
sheet['A1'].fill = PatternFill(start_color='AAEE22', end_color='AAEE22', fill_type='solid')
sheet['A1'].border = Border(left=Side(border_style='medium', color='FF000000'),
                            right=Side(border_style='medium', color='FF000000'),
                            top=Side(border_style='medium', color='FF000000'),
                            bottom=Side(border_style='medium', color='FF000000'))
sheet['B1'] = 'SAE-GW'
sheet['B1'].font = Font(name='Courrier', size=12, bold=False)
sheet['B1'].alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False,
                                  shrink_to_fit=False, indent=0)
sheet['B1'].fill = PatternFill(start_color='FFEE11', end_color='FFEE11', fill_type='solid')
sheet['B1'].border = Border(left=Side(border_style='medium', color='FF000000'),
                            right=Side(border_style='medium', color='FF000000'),
                            top=Side(border_style='medium', color='FF000000'),
                            bottom=Side(border_style='medium', color='FF000000'))
f = open('./Prints_UGW/ip-pool/IPs_publicos_ip-range-list.txt')
lines = f.readlines()
f.close()
for line in lines:
    rede = find_before(line, '-')
    elemento = find_after(line, '-').replace('\n', '')
    if rede == 'ERROR': rede = line.replace('\n', '')
    if elemento == 'ERROR': elemento = 'livre'
    ws.append([rede, elemento])
datahora = datetime.datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d_%H%M%S')
arquivo = './Prints_UGW/' + datahora + ' - Controle_IP.xlsx'
wb.save(filename=arquivo)
# --------------------------------

print
print 'Arquivo salvo: ' + arquivo
print
confirma = raw_input("Tecle algo para terminar...")

print "Concluido..."

# --------------------------------
