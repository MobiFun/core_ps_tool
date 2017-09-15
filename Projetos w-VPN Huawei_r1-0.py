# -*- coding: utf-8 -*-


import datetime
import time
import os
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Style, Font, PatternFill, Border, Side, Alignment, Protection


# -----------------------------------------------------------------
# !/usr/bin/env python

# -----------------------------------------------------------------
def desmembrar_IP(IP):
    octetos = []
    start = 0
    end = IP.index(".", start)
    octeto1 = IP[start:end]
    start = end + 1
    end = IP.index(".", start)
    octeto2 = IP[start:end]
    start = end + 1
    end = IP.index(".", start)
    octeto3 = IP[start:end]
    start = end + 1
    try:
        end = IP.index("/", start)
        octeto4 = IP[start:end]
        start = end + 1
        mask = IP[start:]
    except:
        octeto4 = IP[start:]
        mask = 32

    octetos.append(octeto1)
    octetos.append(octeto2)
    octetos.append(octeto3)
    octetos.append(octeto4)
    octetos.append(mask)
    return (octetos)


# -----------------------------------------------------------------


# -----------------------------------------------------------------
def calcula_wild(mask):
    resp = ""
    zeros = int(mask) / 8
    resto = int(mask) % 8
    nonzero = (2 ** (8 - resto)) - 1
    for a in range(zeros):
        resp += "0."
    resp += str(nonzero)
    for b in range(4 - zeros - 1):
        resp += ".255"
    if mask == "32":
        resp = "0.0.0.0"
    return (resp)


# -----------------------------------------------------------------





















# -----------------------------------------------------------------
# Criar ACL-node
def criarACLnode(b, aba):
    resp = ""

    # xlsx
    ws = wb.get_sheet_by_name(aba)
    ws['B' + str(b)] = '#Configuracao da ACL correspondente aos Filtros Layer 3/Layer 4 como redirecionamento'
    c = 1
    for a in range(1, len(rede_cli) + 1):
        ws['C' + str(b + c)] = 'acl-node aln_wvpn_' + cliente + '_' + str(a) + ' filter f_wvpn_' + cliente + '_' + str(
            a) + ' gate pass'
        c += 1
    ws['C' + str(b + c)] = 'acl-node aln_deny_' + cliente + '_all filter f_deny_' + cliente + '_all gate discard'
    c += 3

    for x in range(b, b + c):
        ws['B' + str(x)].border = Border(left=Side(border_style='double', color='FF000000'))
    b += c
    return (b)


# -----------------------------------------------------------------


# -----------------------------------------------------------------
# Criar acl match order
def criarMatchOrder(a, aba):
    resp = ""

    # xlsx
    ws = wb.get_sheet_by_name(aba)
    ws['B' + str(a)] = '#Especificacao do nome do ACL e da ordem correspondente'
    ws['C' + str(a + 1)] = 'acl acl_wvpn_' + cliente + ' match-order auto'

    for x in range(a, a + 4):
        ws['B' + str(x)].border = Border(left=Side(border_style='double', color='FF000000'))
    a += 4
    return (a)


# -----------------------------------------------------------------


# -----------------------------------------------------------------
# Criar acl-node-bindig
def criarACLnodeBinding(b, aba):
    resp = ""

    # xlsx
    ws = wb.get_sheet_by_name(aba)
    ws['B' + str(b)] = '#Ligacao dos nodes da ACL com a ACL'
    c = 1
    for a in range(1, len(rede_cli) + 1):
        ws['C' + str(b + c)] = 'acl-node-binding acl acl_wvpn_' + cliente + ' acl-node aln_wvpn_' + cliente + '_' + str(
            a)
        c += 1
    ws['C' + str(b + c)] = 'acl-node-binding acl acl_wvpn_' + cliente + ' acl-node aln_deny_' + cliente + '_all'
    ws['C' + str(b + c + 1)] = 'refresh-service'
    ws['C' + str(b + c + 2)] = 'quit'
    c += 5

    for x in range(b, b + c):
        ws['B' + str(x)].border = Border(left=Side(border_style='double', color='FF000000'))
    b += c
    return (b)


# -----------------------------------------------------------------


# -----------------------------------------------------------------
# Criar Tunnel GRE
def criarTunnelGRE(a, aba, peer, IP_tunnel):
    resp = ""

    # xlsx
    ws = wb.get_sheet_by_name(aba)
    ws['B' + str(a)] = '# Criar Interface tunnel'
    ws['C' + str(a + 1)] = 'interface Tunnel' + str(int(APNID))
    ws['C' + str(a + 2)] = ' description tunnel ' + vrf
    ws['C' + str(a + 3)] = ' ip binding vpn-instance ' + vrf
    ws['C' + str(a + 4)] = ' ip address ' + IP_tunnel + ' 255.255.255.252'
    ws['C' + str(a + 5)] = ' tunnel-protocol gre'
    ws['C' + str(a + 6)] = ' source LoopBack1'
    ws['C' + str(a + 7)] = ' destination vpn-instance GI-CORP ' + peer

    for x in range(a, a + 10):
        ws['B' + str(x)].border = Border(left=Side(border_style='double', color='FF000000'))
    a += 10
    return (a)


# -----------------------------------------------------------------


# -----------------------------------------------------------------
# Criar acl-binding
def criarACLbinding(a, aba):
    resp = ""

    # xlsx
    ws = wb.get_sheet_by_name(aba)
    ws['C' + str(a)] = ' acl-binding direction up-in acl acl_wvpn_' + cliente
    ws['C' + str(a + 1)] = ' acl-binding direction down-in acl acl_wvpn_' + cliente

    for x in range(a, a + 4):
        ws['B' + str(x)].border = Border(left=Side(border_style='double', color='FF000000'))
    a += 4
    return (a)


# -----------------------------------------------------------------


# -----------------------------------------------------------------
# Criar acl-binding
def cabecalho(a, aba, elemento, cor):
    ws = wb.get_sheet_by_name(aba)
    ws['A' + str(a)] = ''
    ws.column_dimensions['A'].width = 3.0
    for x in range(1, 400):
        for b in range(1, 30):
            ws.cell(row=x, column=b).fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')
    ws['B' + str(a)].font = Font(name='Courrier', size=14, bold=True)
    ws['B' + str(a)] = 'Network Engineering - Packet Switching'
    ws['B' + str(a)].border = Border(left=Side(border_style='thick', color='FF000000'))
    ws['B' + str(a + 1)] = 'Projeto:'
    ws['B' + str(a + 1)].border = Border(left=Side(border_style='thick', color='FF000000'))
    ws['C' + str(a + 1)] = 'w-VPN ' + cliente + ' - APN ' + APN
    ws['B' + str(a + 2)] = "Versão:"
    ws['B' + str(a + 2)].border = Border(left=Side(border_style='thick', color='FF000000'))
    ws['C' + str(a + 2)] = '1.0'
    ws['B' + str(a + 3)] = 'Data:'
    ws['B' + str(a + 3)].border = Border(left=Side(border_style='thick', color='FF000000'),
                                         bottom=Side(border_style='thick', color='FF000000'))
    ws['C' + str(a + 3)] = datetime.datetime.fromtimestamp(time.time()).strftime('%d-%m-%Y')
    colunas = 'CDEFGHIJ'
    for x in colunas:
        ws[x + str(a + 3)].border = Border(bottom=Side(border_style='thick', color='FF000000'))
    # ws['B'+str(a+6)] = 'Elemento:'
    # ws['B'+str(a+6)].border = Border(bottom=Side(border_style='double',color='FF000000'))
    ws['B' + str(a + 6)].font = Font(name='Courrier', size=14, bold=True)
    ws['B' + str(a + 6)] = elemento
    """
    ws['B'+str(a+6)].border = Border(bottom=Side(border_style='double',color='FF000000'))
    ws['C'+str(a+6)].border = Border(bottom=Side(border_style='double',color='FF000000'))
    ws['D'+str(a+6)].border = Border(bottom=Side(border_style='double',color='FF000000'))
    ws['E'+str(a+6)].border = Border(bottom=Side(border_style='double',color='FF000000'))
    ws['F'+str(a+6)].border = Border(bottom=Side(border_style='double',color='FF000000'))
    """
    colunas = 'BCDEFGHIJ'
    for x in colunas:
        ws[x + str(a + 6)].border = Border(bottom=Side(border_style='double', color='FF000000'))
    ws['B' + str(a + 7)].border = Border(left=Side(border_style='double', color='FF000000'))
    ws['B' + str(a + 8)].border = Border(left=Side(border_style='double', color='FF000000'))
    a += 9

    return (a)


# -----------------------------------------------------------------






# ok
# -----------------------------------------------------------------
def criar_VRF():  # VRF, APNID
    saida = u"Configuração PGWs (GPBSA1/GPBHE1)\r\n"
    saida += "-------------------------------------------------------------------------------------------\r\n\r\n"
    saida += "# Criar VRF\r\n"
    saida += "\tsystem-view\r\n"
    saida += "\tip vpn-instance " + VRF + "\r\n"
    saida += "\t ipv4-family\r\n"
    saida += "\t  route-distinguisher 1:" + str(APNID) + "\r\n"
    saida += "\t  quit\r\n"
    saida += "\t quit\r\n"
    return (saida)


# -----------------------------------------------------------------


# -----------------------------------------------------------------
def criar_ACL():  # APNID, VRF, GPBSA_range, GPBHE_range
    saida = u"\r\n\r\n# Criar ACL (Domínio de encriptação do túnel IPsec)\r\n\r\n"
    saida += "    ### Configurar em GPBSA1\r\n"
    if len(str(APNID)) == 2: lastro = "0"
    if len(str(APNID)) == 1: lastro = "00"
    if len(str(APNID)) == 3: lastro = ""
    saida += "    +--\tacl number 6" + lastro + str(APNID) + "\r\n"
    saida += "    |\t description acl " + VRF + "\r\n"
    rule = 10
    for a in range(len(rede_cli)):
        octetos_local = desmembrar_IP(GPBSA_range)
        wildcard_local = calcula_wild(octetos_local[4])
        rede_wild_local = octetos_local[0] + '.' + octetos_local[1] + '.' + octetos_local[2] + '.' + octetos_local[
            3] + ' ' + wildcard_local
        if rede_cli[a].lower() != "any":
            octetos_remoto = desmembrar_IP(rede_cli[a])
            wildcard_remoto = calcula_wild(octetos_remoto[4])
            rede_wild_remoto = octetos_remoto[0] + '.' + octetos_remoto[1] + '.' + octetos_remoto[2] + '.' + \
                               octetos_remoto[3] + ' ' + wildcard_remoto
        else:
            rede_wild_remoto = "any"
        saida += "    |\t rule " + str(
            rule) + " permit ip source ip-address " + rede_wild_local + " destination ip-address " + rede_wild_remoto + "\r\n"
        rule += 10
    saida += "    +--\t quit\r\n"

    saida += u"\r\n\r\n"
    saida += "    ### Configurar em GPBHE1\r\n"
    saida += "    +--\tacl number 6" + lastro + str(APNID) + "\r\n"
    saida += "    |\t description acl " + VRF + "\r\n"
    rule = 10
    for a in range(len(rede_cli)):
        octetos_local = desmembrar_IP(GPBHE_range)
        wildcard_local = calcula_wild(octetos_local[4])
        rede_wild_local = octetos_local[0] + '.' + octetos_local[1] + '.' + octetos_local[2] + '.' + octetos_local[
            3] + ' ' + wildcard_local
        if rede_cli[a].lower() != "any":
            octetos_remoto = desmembrar_IP(rede_cli[a])
            wildcard_remoto = calcula_wild(octetos_remoto[4])
            rede_wild_remoto = octetos_remoto[0] + '.' + octetos_remoto[1] + '.' + octetos_remoto[2] + '.' + \
                               octetos_remoto[3] + ' ' + wildcard_remoto
        else:
            rede_wild_remoto = "any"
        saida += "    |\t rule " + str(
            rule) + " permit ip source ip-address " + rede_wild_local + " destination ip-address " + rede_wild_remoto + "\r\n"
        rule += 10
    saida += "    +--\t quit\r\n"
    return (saida)


# -----------------------------------------------------------------


# -----------------------------------------------------------------
# Criar IKE peer
def criar_IKE_peer():  # GPBSA_peer, GPBHE_peer, VRF, policy, GPBSA_peer_bkp, GPBHE_peer_bkp
    prop = {'aes128_sha1': '1', '3des_sha1': '2', 'aes256_sha1': '3', 'aes128_md5': '5', '3des_md5': '6',
            'aes256_md5': '7'}
    saida = u"\r\n\r\n# Criar IKE peer\r\n"
    if GPBSA_peer != GPBHE_peer:
        saida += "\r\n    ### Configurar em GPBSA1\r\n"
        saida += "    +--\tike peer " + VRF + "\r\n"
        saida += "    |\t pre-shared-key simple " + PSK + "\r\n"
        saida += "    |\t ike-proposal " + prop[policy] + "\r\n"
        saida += "    |\t local-id-type ip check disable\r\n"
        saida += "    |\t remote-address " + GPBSA_peer + "\r\n"
        if GPBSA_peer_bkp != None:
            saida += "    |\t ike dpd retry-interval 10\r\n"
        if GPBSA_peer_bkp != None:
            saida += "    |\t quit\r\n"
            saida += "    |\r\n"
            saida += "    |\tike peer " + VRF + "_bkp\r\n"
            saida += "    |\t pre-shared-key simple " + PSK + "\r\n"
            saida += "    |\t ike-proposal " + prop[policy] + "\r\n"
            saida += "    |\t local-id-type ip check disable\r\n"
            saida += "    |\t remote-address " + GPBSA_peer_bkp + "\r\n"
            saida += "    |\t ike dpd retry-interval 10\r\n"
        saida += "    +--\t quit\r\n"

        saida += u"\r\n\r\n"
        saida += "    ### Configurar em GPBHE1\r\n"
        saida += "    +--\tike peer " + VRF + "\r\n"
        saida += "    |\t pre-shared-key simple " + PSK + "\r\n"
        saida += "    |\t ike-proposal " + prop[policy] + "\r\n"
        saida += "    |\t local-id-type ip check disable\r\n"
        saida += "    |\t remote-address " + GPBHE_peer + "\r\n"
        if GPBHE_peer_bkp != None:
            saida += "    |\t ike dpd retry-interval 10\r\n"
        if GPBHE_peer_bkp != None:
            saida += "    |\t quit\r\n"
            saida += "    |\r\n"
            saida += "    |\tike peer " + VRF + "_bkp\r\n"
            saida += "    |\t pre-shared-key simple " + PSK + "\r\n"
            saida += "    |\t ike-proposal " + prop[policy] + "\r\n"
            saida += "    |\t local-id-type ip check disable\r\n"
            saida += "    |\t remote-address " + GPBHE_peer_bkp + "\r\n"
            saida += "    |\t ike dpd retry-interval 10\r\n"
        saida += "    +--\t quit\r\n"

    else:
        saida += "\tike peer " + VRF + "\r\n"
        saida += "\t pre-shared-key simple " + PSK + "\r\n"
        saida += "\t ike-proposal " + prop[policy] + "\r\n"
        saida += "\t local-id-type ip check disable\r\n"
        saida += "\t remote-address " + GPBSA_peer + "\r\n"
        if GPBSA_peer_bkp != None:
            saida += "\t ike dpd retry-interval 10\r\n"
        saida += "\t quit\r\n"

        if GPBSA_peer_bkp != "":
            saida += "\r\n"
            saida += "\tike peer " + VRF + "_bkp\r\n"
            saida += "\t pre-shared-key simple " + PSK + "\r\n"
            saida += "\t ike-proposal " + prop[policy] + "\r\n"
            saida += "\t local-id-type ip check disable\r\n"
            saida += "\t remote-address " + GPBSA_peer_bkp + "\r\n"
            if GPBSA_peer_bkp != None:
                saida += "\t ike dpd retry-interval 10\r\n"
            saida += "\t quit\r\n"
    return (saida)


# -----------------------------------------------------------------


# -----------------------------------------------------------------
# Criar IPsec Policy
def criar_IPSEC_policy():  # VRF, APNID, GPBSA_peer_bkp, policy
    saida = "\r\n\r\n# Criar IPsec Policy\r\n"
    saida += "\tipsec policy " + VRF + " 1 isakmp\r\n"
    if len(str(APNID)) == 2: lastro = "0"
    if len(str(APNID)) == 1: lastro = "00"
    if len(str(APNID)) == 3: lastro = ""
    saida += "    \t security acl 6" + lastro + str(APNID) + "\r\n"
    saida += "\t pfs dh-group2\r\n"
    saida += "\t ike-peer " + VRF + "\r\n"
    if GPBSA_peer_bkp != None:
        saida += "\t ike-peer " + VRF + "_bkp 2\r\n"
    saida += "\t proposal " + policy + "\r\n"
    saida += "\t sa duration time-based 3600\r\n"
    saida += "\t quit\r\n"
    return (saida)


# -----------------------------------------------------------------


# -----------------------------------------------------------------
# Criar interface tunnel IPSEC
def criar_Interface_tunnel_IPSEC():  # APNID, VRF, IP_tunnel_BSA, IP_tunnel_BHE
    saida = "\r\n\r\n# Criar Interface tunnel\r\n"
    saida += "\r\n    ### Configurar em GPBSA1\r\n"
    saida += "    +--\tinterface Tunnel" + str(APNID) + "\r\n"
    saida += "    |\t description tunnel " + VRF + "\r\n"
    saida += "    |\t ip binding vpn-instance " + VRF + "\r\n"
    saida += "    |\t ip address " + IP_tunnel_BSA + " 255.255.255.255\r\n"
    saida += "    |\t tunnel-protocol ipsec\r\n"
    saida += "    |\t source LoopBack2\r\n"
    saida += "    |\t ipsec policy " + VRF + "\r\n"
    saida += "    +--\t quit\r\n"
    # -------------------
    saida += "\r\n"
    saida += "\r\n    ### Configurar em GPBHE1\r\n"
    saida += "    +--\tinterface Tunnel" + str(APNID) + "\r\n"
    saida += "    |\t description tunnel " + VRF + "\r\n"
    saida += "    |\t ip binding vpn-instance " + VRF + "\r\n"
    saida += "    |\t ip address " + IP_tunnel_BHE + " 255.255.255.255\r\n"
    saida += "    |\t tunnel-protocol ipsec\r\n"
    saida += "    |\t source LoopBack2\r\n"
    saida += "    |\t ipsec policy " + VRF + "\r\n"
    saida += "    +--\t quit\r\n"
    return (saida)


# -----------------------------------------------------------------


# -----------------------------------------------------------------
# Criar IP Pool
def criar_IP_pool():  # VRF, Addr_aloc, GPBSA_range, GPBHE_range
    oct3 = {'17': 127, '18': 63, '19': 31, '20': 15, '21': 7, '22': 3, '23': 1, '24': 0}
    saida = "\r\n\r\n# Criar IP Pool\r\n"
    saida += "\r\n    ### Configurar em GPBSA1\r\n"
    if Addr_aloc == "local":
        saida += "    +--\tip pool " + VRF + " local ipv4\r\n"
    else:
        saida += "    +--\tip pool " + VRF + " radius ipv4 group 1\r\n"
    saida += "    |\t vpn-instance " + VRF + "\r\n"
    octetos = desmembrar_IP(GPBSA_range)
    if int(octetos[4]) > 16:
        saida += "    |\t section 1 " + octetos[0] + "." + octetos[1] + "." + octetos[2] + ".2 " + octetos[0] + "." + \
                 octetos[1] + "." + str(int(octetos[2]) + oct3[str(octetos[4])]) + ".255\r\n"
    else:
        if int(octetos[4]) == 16:
            octetos[4] = '17'
            saida += "    |\t section 1 " + octetos[0] + "." + octetos[1] + "." + octetos[2] + ".2 " + octetos[
                0] + "." + octetos[1] + ".127.255\r\n"
            saida += "    |\t section 2 " + octetos[0] + "." + octetos[1] + ".128.0 " + octetos[0] + "." + octetos[
                1] + ".255.255\r\n"
        else:
            if int(octetos[4]) == 15:
                octetos[4] = '17'
                saida += "    |\t section 1 " + octetos[0] + "." + octetos[1] + ".0.2 " + octetos[0] + "." + octetos[
                    1] + ".127.255\r\n"
                saida += "    |\t section 2 " + octetos[0] + "." + octetos[1] + ".128.0 " + octetos[0] + "." + octetos[
                    1] + ".255.255\r\n"
                saida += "    |\t section 3 " + octetos[0] + "." + str(int(octetos[1]) + 1) + ".0.0 " + octetos[
                    0] + "." + str(int(octetos[1]) + 1) + ".127.255\r\n"
                saida += "    |\t section 4 " + octetos[0] + "." + str(int(octetos[1]) + 1) + ".128.0 " + octetos[
                    0] + "." + str(int(octetos[1]) + 1) + ".255.255\r\n"
    if Addr_aloc == "local": saida += "    |\t alarm-report disable\r\n"
    saida += "    +--\t quit\r\n"
    # -------------------
    saida += "\r\n"
    saida += "\r\n    ### Configurar em GPBHE1\r\n"
    if Addr_aloc == "local":
        saida += "    +--\tip pool " + VRF + " local ipv4\r\n"
    else:
        saida += "    +--\tip pool " + VRF + " radius ipv4 group 1\r\n"
    saida += "    |\t vpn-instance " + VRF + "\r\n"
    octetos = desmembrar_IP(GPBHE_range)
    if int(octetos[4]) > 16:
        saida += "    |\t section 1 " + octetos[0] + "." + octetos[1] + "." + octetos[2] + ".2 " + octetos[0] + "." + \
                 octetos[1] + "." + str(int(octetos[2]) + oct3[str(octetos[4])]) + ".255\r\n"
    else:
        if int(octetos[4]) == 16:
            octetos[4] = '17'
            saida += "    |\t section 1 " + octetos[0] + "." + octetos[1] + "." + octetos[2] + ".2 " + octetos[
                0] + "." + octetos[1] + ".127.255\r\n"
            saida += "    |\t section 2 " + octetos[0] + "." + octetos[1] + ".128.0 " + octetos[0] + "." + octetos[
                1] + ".255.255\r\n"
        else:
            if int(octetos[4]) == 15:
                octetos[4] = '17'
                saida += "    |\t section 1 " + octetos[0] + "." + octetos[1] + ".0.2 " + octetos[0] + "." + octetos[
                    1] + ".127.255\r\n"
                saida += "    |\t section 2 " + octetos[0] + "." + octetos[1] + ".128.0 " + octetos[0] + "." + octetos[
                    1] + ".255.255\r\n"
                saida += "    |\t section 3 " + octetos[0] + "." + str(int(octetos[1]) + 1) + ".0.0 " + octetos[
                    0] + "." + str(int(octetos[1]) + 1) + ".127.255\r\n"
                saida += "    |\t section 4 " + octetos[0] + "." + str(int(octetos[1]) + 1) + ".128.0 " + octetos[
                    0] + "." + str(int(octetos[1]) + 1) + ".255.255\r\n"
    if Addr_aloc == "local": saida += "    |\t alarm-report disable\r\n"
    saida += "    +--\t quit\r\n"
    return (saida)


# -----------------------------------------------------------------


# -----------------------------------------------------------------
# Criar Static Route
def criar_Route():  # VRF, APNID
    saida = "\r\n\r\n# Criar static-route\r\n"
    saida += "\tip route-static vpn-instance " + VRF + " 0.0.0.0 0.0.0.0 Tunnel" + str(APNID) + "\r\n"
    return (saida)


# -----------------------------------------------------------------


# -----------------------------------------------------------------
# Criar Interface Giif
def criar_GiIf():  # APNID, VRF, GPBSA_range, GPBHE_range
    # saida = "ip address {octeto0}.{octeto1}".format(octeto0="ze", octeto1="leo")
    saida = "\r\n\r\n# Criar Interface Giif\r\n"
    saida += "\r\n    ### Configurar em GPBSA1\r\n"
    saida += "    +--\tinterface Giif1/0/" + str(APNID) + "\r\n"
    saida += "    |\t ip binding vpn-instance " + VRF + "\r\n"
    octetos = desmembrar_IP(GPBSA_range)
    saida += "    |\t ip address " + octetos[0] + "." + octetos[1] + "." + octetos[2] + ".1 255.255.255.255\r\n"
    saida += "    +--\t quit\r\n"
    # -------------------
    saida += "\r\n"
    saida += "\r\n    ### Configurar em GPBHE1\r\n"
    saida += "    +--\tinterface Giif1/0/" + str(APNID) + "\r\n"
    saida += "    |\t ip binding vpn-instance " + VRF + "\r\n"
    octetos = desmembrar_IP(GPBHE_range)
    saida += "    |\t ip address " + octetos[0] + "." + octetos[1] + "." + octetos[2] + ".1 255.255.255.255\r\n"
    saida += "    +--\t quit\r\n"
    return (saida)


# -----------------------------------------------------------------


# -----------------------------------------------------------------
# Criar Servidor RADIUS
def criar_Radius_server_group():  # GPBSA_RAD1, GPBHE_RAD1, GPBSA_RAD2, GPBHE_RAD2, VRF, RAD_Acct, RAD_Auth, RAD_NAS_secret, RAD_backup_mode, APNID
    saida = "\r\n\r\n# Criar Servidor Radius\r\n"
    if (GPBSA_RAD1 != GPBHE_RAD1) or (GPBSA_RAD2 != GPBHE_RAD2):
        saida += "\r\n    ### Configurar em GPBSA1\r\n"
        saida += "    +--\taccess-view\r\n"
        saida += "    |\t radius-server group " + VRF + "\r\n"
        if RAD_Acct != None: saida += "    |\t  radius-server accounting " + GPBSA_RAD1 + " port 1813 vpn " + VRF + " key " + RAD_NAS_secret + "\r\n"
        if GPBSA_RAD2 != None and RAD_Acct != None: saida += "    |\t  radius-server accounting " + GPBSA_RAD2 + " port 1813 vpn " + VRF + " key " + RAD_NAS_secret + " " + RAD_backup_mode + "\r\n"
        if tipo == 'Sem Tunel': saida += "    |\t  radius-server accounting 189.40.224.20 port 1813 vpn Gi key ISPPRV carbon-copy\r\n"
        if RAD_Auth != None: saida += "    |\t  radius-server authentication " + GPBSA_RAD1 + " port 1812 vpn " + VRF + " key " + RAD_NAS_secret + "\r\n"
        if GPBSA_RAD2 != None and RAD_Auth != None: saida += "    |\t  radius-server authentication " + GPBSA_RAD2 + " port 1812 vpn " + VRF + " key " + RAD_NAS_secret + " secondary\r\n"
        saida += "    |\t  radius-server mode backup\r\n"
        saida += "    |\t  radius-server authentication retransmit 3 timeout 3 response-timeout 12 deadtime 180\r\n"
        saida += "    |\t  radius-server accounting retransmit 3 timeout 3 response-timeout 12 deadtime 180\r\n"
        saida += "    |\t  radius-server acct-attribute nas-id enable sys-name\r\n"
        saida += "    |\t  radius-server acct-3gppvsa 3gpp enable\r\n"
        if RAD_Acct != None: saida += "    |\t  client-ip acct interface Giif1/0/" + str(APNID) + "\r\n"
        if RAD_Auth != None: saida += "    |\t  client-ip auth interface Giif1/0/" + str(APNID) + "\r\n"
        if tipo == 'Sem Tunel': saida += "    |\t  client-ip acct interface Giif1/0/0\r\n"
        saida += "    |\t  radius-server auth-attribute\r\n"
        saida += "    |\t  radius-server auth-3gppvsa 3gpp enable\r\n"
        saida += "    |\t  radius-server accept-attribute\r\n"
        saida += "    |\t  quit\r\n"
        saida += "    +--\t quit\r\n"
        # -------------------
        saida += "\r\n"
        saida += "\r\n    ### Configurar em GPBHE1\r\n"
        saida += "    +--\taccess-view\r\n"
        saida += "    |\t radius-server group " + VRF + "\r\n"
        if RAD_Acct != None: saida += "    |\t  radius-server accounting " + GPBHE_RAD1 + " port 1813 vpn " + VRF + " key " + RAD_NAS_secret + "\r\n"
        if GPBHE_RAD2 != None and RAD_Acct != None: saida += "    |\t  radius-server accounting " + GPBHE_RAD2 + " port 1813 vpn " + VRF + " key " + RAD_NAS_secret + " " + RAD_backup_mode + "\r\n"
        if tipo == 'Sem Tunel': saida += "    |\t  radius-server accounting 189.40.224.20 port 1813 vpn Gi key ISPPRV carbon-copy\r\n"
        if RAD_Auth != None: saida += "    |\t  radius-server authentication " + GPBHE_RAD1 + " port 1812 vpn " + VRF + " key " + RAD_NAS_secret + "\r\n"
        if GPBHE_RAD2 != None and RAD_Auth != None: saida += "    |\t  radius-server authentication " + GPBHE_RAD2 + " port 1812 vpn " + VRF + " key " + RAD_NAS_secret + " secondary\r\n"
        saida += "    |\t  radius-server mode backup\r\n"
        saida += "    |\t  radius-server authentication retransmit 3 timeout 3 response-timeout 12 deadtime 180\r\n"
        saida += "    |\t  radius-server accounting retransmit 3 timeout 3 response-timeout 12 deadtime 180\r\n"
        saida += "    |\t  radius-server acct-attribute nas-id enable sys-name\r\n"
        saida += "    |\t  radius-server acct-3gppvsa 3gpp enable\r\n"
        if RAD_Acct != None: saida += "    |\t  client-ip acct interface Giif1/0/" + str(APNID) + "\r\n"
        if RAD_Auth != None: saida += "    |\t  client-ip auth interface Giif1/0/" + str(APNID) + "\r\n"
        if tipo == 'Sem Tunel': saida += "    |\t  client-ip acct interface Giif1/0/0\r\n"
        saida += "    |\t  radius-server auth-attribute\r\n"
        saida += "    |\t  radius-server auth-3gppvsa 3gpp enable\r\n"
        saida += "    |\t  radius-server accept-attribute\r\n"
        saida += "    |\t  quit\r\n"
        saida += "    +--\t quit\r\n"
    else:
        saida += "\taccess-view\r\n"
        saida += "\t radius-server group " + VRF + "\r\n"
        if RAD_Acct != None: saida += "\t  radius-server accounting " + GPBSA_RAD1 + " port 1813 vpn " + VRF + " key " + RAD_NAS_secret + "\r\n"
        if GPBSA_RAD2 != None and RAD_Acct != None: saida += "\t  radius-server accounting " + GPBSA_RAD2 + " port 1813 vpn " + VRF + " key " + RAD_NAS_secret + " " + RAD_backup_mode + "\r\n"
        if RAD_Auth != None: saida += "\t  radius-server authentication " + GPBSA_RAD1 + " port 1812 vpn " + VRF + " key " + RAD_NAS_secret + "\r\n"
        if GPBSA_RAD2 != None and RAD_Auth != None: saida += "\t  radius-server authentication " + GPBSA_RAD2 + " port 1812 vpn " + VRF + " key " + RAD_NAS_secret + " secondary\r\n"
        saida += "\t  radius-server mode backup\r\n"
        saida += "\t  radius-server authentication retransmit 3 timeout 3 response-timeout 12 deadtime 180\r\n"
        saida += "\t  radius-server accounting retransmit 3 timeout 3 response-timeout 12 deadtime 180\r\n"
        saida += "\t  radius-server acct-attribute nas-id enable sys-name\r\n"
        saida += "\t  radius-server acct-3gppvsa 3gpp enable\r\n"
        if RAD_Acct != None: saida += "\t  client-ip acct interface Giif1/0/" + str(APNID) + "\r\n"
        if RAD_Auth != None: saida += "\t  client-ip auth interface Giif1/0/" + str(APNID) + "\r\n"
        saida += "\t  radius-server auth-attribute\r\n"
        saida += "\t  radius-server auth-3gppvsa 3gpp enable\r\n"
        saida += "\t  radius-server accept-attribute\r\n"
        saida += "\t  quit\r\n"
        saida += "\t quit\r\n"
    return (saida)


# -----------------------------------------------------------------


# -----------------------------------------------------------------
# Criar APN
def criar_APN():
    saida = "\r\n\r\n# Criar APN\r\n"
    saida += "\tapn " + APN + "\r\n"
    if tipo == "Sem Tunel":
        saida += "\t vpn-instance Gi\r\n"
    else:
        saida += "\t vpn-instance " + VRF + "\r\n"
    saida += "\t content-awareness disable\r\n"
    saida += "\t service-report switch global\r\n"
    if RAD_Auth == None:
        saida += "\t access-mode transparent-non-authentication\r\n"
    else:
        if auth_type == "MSISDN":
            saida += "\t access-mode non-transparent authentication-mode msisdn authentication-password password pco-priority disable\r\n"
        else:
            saida += "\t access-mode non-transparent authentication-mode pco\r\n"
    saida += "\t framed-route-mode disable\r\n"
    if Addr_aloc == "local":
        saida += "\t address-allocate ipv4 local radius-prior disable ipv6 local radius-prior disable\r\n"
    else:
        saida += "\t address-allocate ipv4 radius ipv6 local radius-prior disable\r\n"
    saida += "\t address-support ipv4 enable ipv6 enable preference ipv4\r\n"
    saida += "\t address-allocate-preference enable\r\n"
    saida += "\t ppp-access authentication disable\r\n"
    saida += "\t ppp-access address-allocate local radius-prior disable\r\n"
    saida += "\t virtual-apn disable\r\n"
    saida += "\t address-inherit enable\r\n"
    saida += "\t apn-restriction disable\r\n"
    saida += "\t remove-domain-name radius disable\r\n"
    saida += "\t remove-domain-name lns disable\r\n"
    saida += "\t roaming-user-access sgw enable visiting-user-access enable\r\n"
    saida += "\t roaming-user-access ggsn-pgw enable visiting-user-access enable\r\n"
    saida += "\t session-timeout disable\r\n"
    saida += "\t idle-timeout disable\r\n"
    saida += "\t static-ip hlr-hss-provided disable route enable all\r\n"
    saida += "\t select-mode-check enable\r\n"
    saida += "\t lock disable\r\n"
    if GPBSA_RAD1 != None:
        # if tipo == "IPsec" or tipo == "GRE" or tipo == "LD":
        saida += "\t radius-server group " + VRF + " copy-interim-update enable\r\n"
    else:
        if tipo == 'Sem Tunel': saida += "\t radius-server group timbrasil copy-interim-update disable\r\n"
    if DNS1 != None and DNS2 != None:
        saida += "\t dns ipv4 primary-ip " + DNS1 + " secondary-ip " + DNS2 + "\r\n"
    else:
        if DNS1 != None and DNS2 == None:
            saida += "\t dns ipv4 primary-ip " + DNS1 + "\r\n"
    if tipo == "Sem Tunel":
        saida += "\t address-pool pool-wvpn\r\n"
    else:
        saida += "\t address-pool " + VRF + "\r\n"
    saida += "\t volume-statistic-mode layer-all\r\n"
    saida += "\t aaa-apn-secondauth disable\r\n"
    saida += "\t apn-type-select perf service cg service aaaacct service aaaauth service ocs service pcrf service header-enrichment service\r\n"
    saida += "\t plmn serving-node-mapping enable\r\n"
    saida += "\t rat sgsn-sgw-mapping enable\r\n"
    saida += "\t multiple-service-mode radius\r\n"
    saida += "\t radius-disconnect disable\r\n"
    saida += "\t cdr-field-binding gcdr cdr pgw-cdr cdr sgw-cdr cdr\r\n"
    saida += "\t offline-charge-binding ggsn cdr pgw cdr sgw cdr\r\n"
    saida += "\t radius acctctrl accounting-start service-trigger enable wait-accounting-response disable response-timeout continue  accounting-stop enable accounting-update enable\r\n"
    saida += "\t serving-gateway charge-method offline enable\r\n"
    saida += "\t pcc-switch enable\r\n"
    saida += "\t pcc-default reporting-level rg metering-method volume\r\n"
    saida += "\t pcrf-group-binding pcrf-group tekelec\r\n"
    if tipo == 'Sem Tunel' and URLs_cli != []:
        saida += "\t user-profile-group-binding wvpn_" + VRF + "\r\n"
    else:
        saida += "\t user-profile-group-binding wvpn_downinit\r\n"
    if tipo == 'Sem Tunel' and URLs_cli == []:
        saida += "\t acl-binding direction up-in acl acl_wvpn_" + VRF + "\r\n"
        saida += "\t acl-binding direction down-in acl acl_wvpn_" + VRF + "\r\n"
    saida += "\t tcp-mss ipv4 1400 ipv6 1400\r\n"
    saida += "\t quit\r\n"
    return (saida)


# -----------------------------------------------------------------


# -----------------------------------------------------------------
# Criar Alias APN
def criar_Alias_APN():
    saida = "\r\n\r\n# Criar Alias APN\r\n"
    saida += "\taccess-view\r\n"
    saida += "\t apn-relate apn " + APN + " relate-name " + Alias1_APN + "\r\n"
    if Alias2_APN != None: saida += "\t apn-relate apn " + APN + " relate-name " + Alias2_APN + "\r\n"
    saida += "\t quit\r\n"
    saida += "\tquit\r\n"
    return (saida)


# -----------------------------------------------------------------


# -----------------------------------------------------------------
# Criar DNS
def criar_DNS():
    PGW_IP = {'GPBSA1': '189.40.166.177', 'GPBHE1': '189.40.166.209', 'GPSNE1': '189.40.173.17',
              'GPRJO4': '189.40.172.49'}
    saida = u"\r\n\r\n# Adicionar novos apontamentos: (arquivo: Configuração_DNS_r1.0.txt)\r\n\r\n"
    saida += '{"action": "add", "domain_name":"' + str(APN) + '.mnc002.mcc724.gprs.", "ip_address": "' + PGW_IP[
        PGWs[0]] + '", "ttl":300}' + "\r\n"
    saida += '{"action": "add", "domain_name":"' + str(APN) + '.mnc003.mcc724.gprs.", "ip_address": "' + PGW_IP[
        PGWs[0]] + '", "ttl":300}' + "\r\n"
    saida += '{"action": "add", "domain_name":"' + str(APN) + '.mnc004.mcc724.gprs.", "ip_address": "' + PGW_IP[
        PGWs[0]] + '", "ttl":300}' + "\r\n"
    saida += '{"action": "add", "domain_name":"' + str(APN) + '.mnc002.mcc724.gprs.", "ip_address": "' + PGW_IP[
        PGWs[1]] + '", "ttl":300}' + "\r\n"
    saida += '{"action": "add", "domain_name":"' + str(APN) + '.mnc003.mcc724.gprs.", "ip_address": "' + PGW_IP[
        PGWs[1]] + '", "ttl":300}' + "\r\n"
    saida += '{"action": "add", "domain_name":"' + str(APN) + '.mnc004.mcc724.gprs.", "ip_address": "' + PGW_IP[
        PGWs[1]] + '", "ttl":300}' + "\r\n"
    if Alias1_APN != None:
        saida += '{"action": "add", "domain_name":"' + str(Alias1_APN) + '.mnc002.mcc724.gprs.", "ip_address": "' + \
                 PGW_IP[PGWs[0]] + '", "ttl":300}' + "\r\n"
        saida += '{"action": "add", "domain_name":"' + str(Alias1_APN) + '.mnc003.mcc724.gprs.", "ip_address": "' + \
                 PGW_IP[PGWs[0]] + '", "ttl":300}' + "\r\n"
        saida += '{"action": "add", "domain_name":"' + str(Alias1_APN) + '.mnc004.mcc724.gprs.", "ip_address": "' + \
                 PGW_IP[PGWs[0]] + '", "ttl":300}' + "\r\n"
        saida += '{"action": "add", "domain_name":"' + str(Alias1_APN) + '.mnc002.mcc724.gprs.", "ip_address": "' + \
                 PGW_IP[PGWs[1]] + '", "ttl":300}' + "\r\n"
        saida += '{"action": "add", "domain_name":"' + str(Alias1_APN) + '.mnc003.mcc724.gprs.", "ip_address": "' + \
                 PGW_IP[PGWs[1]] + '", "ttl":300}' + "\r\n"
        saida += '{"action": "add", "domain_name":"' + str(Alias1_APN) + '.mnc004.mcc724.gprs.", "ip_address": "' + \
                 PGW_IP[PGWs[1]] + '", "ttl":300}' + "\r\n"
    if Alias2_APN != None:
        saida += '{"action": "add", "domain_name":"' + str(Alias2_APN) + '.mnc002.mcc724.gprs.", "ip_address": "' + \
                 PGW_IP[PGWs[0]] + '", "ttl":300}' + "\r\n"
        saida += '{"action": "add", "domain_name":"' + str(Alias2_APN) + '.mnc003.mcc724.gprs.", "ip_address": "' + \
                 PGW_IP[PGWs[0]] + '", "ttl":300}' + "\r\n"
        saida += '{"action": "add", "domain_name":"' + str(Alias2_APN) + '.mnc004.mcc724.gprs.", "ip_address": "' + \
                 PGW_IP[PGWs[0]] + '", "ttl":300}' + "\r\n"
        saida += '{"action": "add", "domain_name":"' + str(Alias2_APN) + '.mnc002.mcc724.gprs.", "ip_address": "' + \
                 PGW_IP[PGWs[1]] + '", "ttl":300}' + "\r\n"
        saida += '{"action": "add", "domain_name":"' + str(Alias2_APN) + '.mnc003.mcc724.gprs.", "ip_address": "' + \
                 PGW_IP[PGWs[1]] + '", "ttl":300}' + "\r\n"
        saida += '{"action": "add", "domain_name":"' + str(Alias2_APN) + '.mnc004.mcc724.gprs.", "ip_address": "' + \
                 PGW_IP[PGWs[1]] + '", "ttl":300}' + "\r\n"
    saida += '{"action": "add", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
             PGWs[0] + '.node.epc.mnc002.mcc724.3gppnetwork.org.", "domain_name": "' + str(
        APN) + '.apn.epc.mnc002.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
    saida += '{"action": "add", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
             PGWs[0] + '.node.epc.mnc003.mcc724.3gppnetwork.org.", "domain_name": "' + str(
        APN) + '.apn.epc.mnc003.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
    saida += '{"action": "add", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
             PGWs[0] + '.node.epc.mnc004.mcc724.3gppnetwork.org.", "domain_name": "' + str(
        APN) + '.apn.epc.mnc004.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
    saida += '{"action": "add", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
             PGWs[1] + '.node.epc.mnc002.mcc724.3gppnetwork.org.", "domain_name": "' + str(
        APN) + '.apn.epc.mnc002.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
    saida += '{"action": "add", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
             PGWs[1] + '.node.epc.mnc003.mcc724.3gppnetwork.org.", "domain_name": "' + str(
        APN) + '.apn.epc.mnc003.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
    saida += '{"action": "add", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
             PGWs[1] + '.node.epc.mnc004.mcc724.3gppnetwork.org.", "domain_name": "' + str(
        APN) + '.apn.epc.mnc004.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
    if Alias1_APN != None:
        saida += '{"action": "add", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[0] + '.node.epc.mnc002.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias1_APN) + '.apn.epc.mnc002.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
        saida += '{"action": "add", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[0] + '.node.epc.mnc003.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias1_APN) + '.apn.epc.mnc003.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
        saida += '{"action": "add", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[0] + '.node.epc.mnc004.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias1_APN) + '.apn.epc.mnc004.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
        saida += '{"action": "add", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[1] + '.node.epc.mnc002.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias1_APN) + '.apn.epc.mnc002.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
        saida += '{"action": "add", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[1] + '.node.epc.mnc003.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias1_APN) + '.apn.epc.mnc003.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
        saida += '{"action": "add", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[1] + '.node.epc.mnc004.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias1_APN) + '.apn.epc.mnc004.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
    if Alias2_APN != None:
        saida += '{"action": "add", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[0] + '.node.epc.mnc002.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias2_APN) + '.apn.epc.mnc002.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
        saida += '{"action": "add", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[0] + '.node.epc.mnc003.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias2_APN) + '.apn.epc.mnc003.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
        saida += '{"action": "add", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[0] + '.node.epc.mnc004.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias2_APN) + '.apn.epc.mnc004.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
        saida += '{"action": "add", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[1] + '.node.epc.mnc002.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias2_APN) + '.apn.epc.mnc002.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
        saida += '{"action": "add", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[1] + '.node.epc.mnc003.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias2_APN) + '.apn.epc.mnc003.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
        saida += '{"action": "add", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[1] + '.node.epc.mnc004.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias2_APN) + '.apn.epc.mnc004.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"

    saida += "\r\n\r\n# Fallback - remover os apontamentos: (arquivo: Fallback_DNS_r1.0.txt)\r\n\r\n"
    saida += '{"action": "remove", "domain_name":"' + str(APN) + '.mnc002.mcc724.gprs.", "ip_address": "' + str(
        PGW_IP[PGWs[0]]) + '", "ttl":300}' + "\r\n"
    saida += '{"action": "remove", "domain_name":"' + str(APN) + '.mnc003.mcc724.gprs.", "ip_address": "' + str(
        PGW_IP[PGWs[0]]) + '", "ttl":300}' + "\r\n"
    saida += '{"action": "remove", "domain_name":"' + str(APN) + '.mnc004.mcc724.gprs.", "ip_address": "' + str(
        PGW_IP[PGWs[0]]) + '", "ttl":300}' + "\r\n"
    saida += '{"action": "remove", "domain_name":"' + str(APN) + '.mnc002.mcc724.gprs.", "ip_address": "' + str(
        PGW_IP[PGWs[1]]) + '", "ttl":300}' + "\r\n"
    saida += '{"action": "remove", "domain_name":"' + str(APN) + '.mnc003.mcc724.gprs.", "ip_address": "' + str(
        PGW_IP[PGWs[1]]) + '", "ttl":300}' + "\r\n"
    saida += '{"action": "remove", "domain_name":"' + str(APN) + '.mnc004.mcc724.gprs.", "ip_address": "' + str(
        PGW_IP[PGWs[1]]) + '", "ttl":300}' + "\r\n"
    if Alias1_APN != None:
        saida += '{"action": "remove", "domain_name":"' + str(
            Alias1_APN) + '.mnc002.mcc724.gprs.", "ip_address": "' + str(PGW_IP[PGWs[0]]) + '", "ttl":300}' + "\r\n"
        saida += '{"action": "remove", "domain_name":"' + str(
            Alias1_APN) + '.mnc003.mcc724.gprs.", "ip_address": "' + str(PGW_IP[PGWs[0]]) + '", "ttl":300}' + "\r\n"
        saida += '{"action": "remove", "domain_name":"' + str(
            Alias1_APN) + '.mnc004.mcc724.gprs.", "ip_address": "' + str(PGW_IP[PGWs[0]]) + '", "ttl":300}' + "\r\n"
        saida += '{"action": "remove", "domain_name":"' + str(
            Alias1_APN) + '.mnc002.mcc724.gprs.", "ip_address": "' + str(PGW_IP[PGWs[1]]) + '", "ttl":300}' + "\r\n"
        saida += '{"action": "remove", "domain_name":"' + str(
            Alias1_APN) + '.mnc003.mcc724.gprs.", "ip_address": "' + str(PGW_IP[PGWs[1]]) + '", "ttl":300}' + "\r\n"
        saida += '{"action": "remove", "domain_name":"' + str(
            Alias1_APN) + '.mnc004.mcc724.gprs.", "ip_address": "' + str(PGW_IP[PGWs[1]]) + '", "ttl":300}' + "\r\n"
    if Alias2_APN != None:
        saida += '{"action": "remove", "domain_name":"' + str(
            Alias2_APN) + '.mnc002.mcc724.gprs.", "ip_address": "' + str(PGW_IP[PGWs[0]]) + '", "ttl":300}' + "\r\n"
        saida += '{"action": "remove", "domain_name":"' + str(
            Alias2_APN) + '.mnc003.mcc724.gprs.", "ip_address": "' + str(PGW_IP[PGWs[0]]) + '", "ttl":300}' + "\r\n"
        saida += '{"action": "remove", "domain_name":"' + str(
            Alias2_APN) + '.mnc004.mcc724.gprs.", "ip_address": "' + str(PGW_IP[PGWs[0]]) + '", "ttl":300}' + "\r\n"
        saida += '{"action": "remove", "domain_name":"' + str(
            Alias2_APN) + '.mnc002.mcc724.gprs.", "ip_address": "' + str(PGW_IP[PGWs[1]]) + '", "ttl":300}' + "\r\n"
        saida += '{"action": "remove", "domain_name":"' + str(
            Alias2_APN) + '.mnc003.mcc724.gprs.", "ip_address": "' + str(PGW_IP[PGWs[1]]) + '", "ttl":300}' + "\r\n"
        saida += '{"action": "remove", "domain_name":"' + str(
            Alias2_APN) + '.mnc004.mcc724.gprs.", "ip_address": "' + str(PGW_IP[PGWs[1]]) + '", "ttl":300}' + "\r\n"
    saida += '{"action": "remove", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
             PGWs[0] + '.node.epc.mnc002.mcc724.3gppnetwork.org.", "domain_name": "' + str(
        APN) + '.apn.epc.mnc002.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
    saida += '{"action": "remove", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
             PGWs[0] + '.node.epc.mnc003.mcc724.3gppnetwork.org.", "domain_name": "' + str(
        APN) + '.apn.epc.mnc003.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
    saida += '{"action": "remove", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
             PGWs[0] + '.node.epc.mnc004.mcc724.3gppnetwork.org.", "domain_name": "' + str(
        APN) + '.apn.epc.mnc004.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
    saida += '{"action": "remove", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
             PGWs[1] + '.node.epc.mnc002.mcc724.3gppnetwork.org.", "domain_name": "' + str(
        APN) + '.apn.epc.mnc002.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
    saida += '{"action": "remove", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
             PGWs[1] + '.node.epc.mnc003.mcc724.3gppnetwork.org.", "domain_name": "' + str(
        APN) + '.apn.epc.mnc003.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
    saida += '{"action": "remove", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
             PGWs[1] + '.node.epc.mnc004.mcc724.3gppnetwork.org.", "domain_name": "' + str(
        APN) + '.apn.epc.mnc004.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
    if Alias1_APN != None:
        saida += '{"action": "remove", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[0] + '.node.epc.mnc002.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias1_APN) + '.apn.epc.mnc002.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
        saida += '{"action": "remove", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[0] + '.node.epc.mnc003.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias1_APN) + '.apn.epc.mnc003.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
        saida += '{"action": "remove", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[0] + '.node.epc.mnc004.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias1_APN) + '.apn.epc.mnc004.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
        saida += '{"action": "remove", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[1] + '.node.epc.mnc002.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias1_APN) + '.apn.epc.mnc002.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
        saida += '{"action": "remove", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[1] + '.node.epc.mnc003.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias1_APN) + '.apn.epc.mnc003.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
        saida += '{"action": "remove", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[1] + '.node.epc.mnc004.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias1_APN) + '.apn.epc.mnc004.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
    if Alias2_APN != None:
        saida += '{"action": "remove", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[0] + '.node.epc.mnc002.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias2_APN) + '.apn.epc.mnc002.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
        saida += '{"action": "remove", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[0] + '.node.epc.mnc003.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias2_APN) + '.apn.epc.mnc003.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
        saida += '{"action": "remove", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[0] + '.node.epc.mnc004.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias2_APN) + '.apn.epc.mnc004.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
        saida += '{"action": "remove", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[1] + '.node.epc.mnc002.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias2_APN) + '.apn.epc.mnc002.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
        saida += '{"action": "remove", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[1] + '.node.epc.mnc003.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias2_APN) + '.apn.epc.mnc003.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
        saida += '{"action": "remove", "flags": "a", "preference": 10, "service": "x-3gpp-pgw:x-s5-gtp:x-gn:x-s8-gtp:x-gp", "ttl": 300, "regexp": "\\"\\"", "replacement": "topoff.vip-s5.' + \
                 PGWs[1] + '.node.epc.mnc004.mcc724.3gppnetwork.org.", "domain_name": "' + str(
            Alias2_APN) + '.apn.epc.mnc004.mcc724.3gppnetwork.org.", "order": 10}' + "\r\n"
    return (saida)


# -----------------------------------------------------------------




# -----------------------------------------------------------------
# Criar Interface tunnel GRE
def criar_Interface_tunnel_GRE():
    saida = "\r\n\r\n# Criar Interface tunnel\r\n"
    saida += "\r\n    ### Configurar em GPBSA1\r\n"
    saida += "    +--\tinterface Tunnel" + str(APNID) + "\r\n"
    saida += "    |\t description tunnel " + VRF + "\r\n"
    saida += "    |\t ip binding vpn-instance " + VRF + "\r\n"
    saida += "    |\t ip address " + IP_tunnel_BSA + " 255.255.255.252\r\n"
    saida += "    |\t tunnel-protocol gre\r\n"
    saida += "    |\t source LoopBack1\r\n"
    saida += "    |\t destination vpn-instance GI-CORP " + GPBSA_peer + "\r\n"
    saida += "    +--\t quit\r\n"
    # -------------------
    saida += "\r\n"
    saida += "\r\n    ### Configurar em GPBHE1\r\n"
    saida += "    +--\tinterface Tunnel" + str(APNID) + "\r\n"
    saida += "    |\t description tunnel " + VRF + "\r\n"
    saida += "    |\t ip binding vpn-instance " + VRF + "\r\n"
    saida += "    |\t ip address " + IP_tunnel_BHE + " 255.255.255.252\r\n"
    saida += "    |\t tunnel-protocol gre\r\n"
    saida += "    |\t source LoopBack1\r\n"
    saida += "    |\t destination vpn-instance GI-CORP " + GPBHE_peer + "\r\n"
    saida += "    +--\t quit\r\n"
    return (saida)


# -----------------------------------------------------------------


# -----------------------------------------------------------------
# Criar filtros
def criar_Filtro():
    F = 0
    FG = 0
    saida = u"Configuração PGWs (GPBSA1/GPBHE1)\r\n"
    saida += "-------------------------------------------------------------------------------------------\r\n\r\n"
    saida += "# Criar Filtros\r\n"
    saida += "\tsystem-view\r\n"
    saida += "\tservice-view\r\n\r\n"

    ## Criar hosts e filters
    if URLs_cli != []:
        FG = 1
        for a in range(len(URLs_cli)):
            saida += "\t host h_wvpn_" + VRF + "_" + str(a) + " domain " + URLs_cli[a] + " sequence 65535\r\n"
        saida += "\r\n"
        for a in range(len(URLs_cli)):
            saida += "\t filter f_wvpn_" + VRF + "_" + str(a) + " l34-protocol any host h_wvpn_" + VRF + "_" + str(
                a) + "\r\n"
            F += 1

    if rede_cli != []:
        for a in range(len(URLs_cli), len(rede_cli) + len(URLs_cli)):
            b = a - len(URLs_cli)
            if rede_cli[b].lower() != "any":
                rede_oct = desmembrar_IP(rede_cli[b])
                rede = rede_oct[0] + "." + rede_oct[1] + "." + rede_oct[2] + "." + rede_oct[3]
                mask = rede_oct[4]
            else:
                rede = u"any"
                mask = ""
            saida += "\t filter f_wvpn_" + VRF + "_" + str(a) + " l34-protocol any server-ip " + rede + " " + str(
                mask) + "\r\n"
            F += 1

    if URLs_cli == []: saida += u"\t filter f_all_block l34-protocol any\t\t\t\t\t\t\t\t(Já existente)\r\n"
    saida += "\r\n"

    ## Criar filter-group se for utilizar "domínio"
    if FG == 1:
        for a in range(F):
            saida += "\t filter-group fg_wvpn_" + VRF + " filter f_wvpn_" + VRF + "_" + str(a) + "\r\n"
        saida += "\r\n"

        ## Criar rule
        saida += "\t rule wvpn_" + VRF + " filter-group fg_wvpn_" + VRF + " service-category-group cg_rg101s7_m8 priority 1135\r\n"
        saida += u"\t rule all_drop filter-group fg_all_drop service-category-group cg_discard priority 49999\t\t\t\t\t\t\t\t(já existente)\r\n\r\n"

        ## Criar user-profile
        saida += "\t user-profile wvpn_" + VRF + "\r\n"
        saida += "\t  rule-binding wvpn_" + VRF + " priority 1135\r\n"
        saida += "\t  rule-binding all_drop priority 49999\r\n"
        saida += "\t  ddos-check disable\r\n"
        saida += "\t  charge-property-binding service-charge-property cp_rg101sid7\r\n"
        saida += "\t  quota-application-action buffer\r\n"
        saida += "\t  quit\r\n\r\n"

        ## Criar user-profile-group
        saida += "\t user-profile-group wvpn_" + VRF + "\r\n"
        saida += "\t  user-profile-binding user-profile-group wvpn_" + VRF + " user-profile wvpn_" + VRF + "\r\n"
        saida += "\t  quit\r\n"
        saida += "\t quit\r\n\r\n"
    else:
        ## Criar acl-node
        for a in range(F):
            saida += "\t acl-node aln_wvpn_" + VRF + "_" + str(a) + " filter f_wvpn_" + VRF + "_" + str(
                a) + " gate pass\r\n"
        saida += u"\t acl-node aln_all_block filter f_all_block gate discard\t\t\t\t\t\t\t\t(já existente)\r\n\r\n"

        saida += "\t acl acl_wvpn_" + VRF + " match-order auto\r\n\r\n"
        ## Criar acl-node-binding
        for a in range(F):
            saida += "\t acl-node-binding acl acl_wvpn_" + VRF + " acl-node aln_wvpn_" + VRF + "_" + str(a) + "\r\n"
        saida += "\t acl-node-binding acl acl_wvpn_" + VRF + " acl-node aln_all_block\r\n"
        saida += "\t quit\r\n"
        saida += "\tquit\r\n\r\n"
    return (saida)


# -----------------------------------------------------------------











###
### Principal
###


path = "./"

##################################################################################
### Abrir planilha datafill (path: ./datafill.xlsx)
try:
    wb = load_workbook(filename=path + 'datafill.xlsx')
except IOError:
    print "Error... planilha nao encontrada"
    quit()

try:
    ws = wb.get_sheet_by_name('datafill')
except KeyError:
    print "Error... - ABA datafill não existe"
    quit()
##################################################################################

##################################################################################
### coleta das infomações
sheet = ws
cliente = sheet['B2'].value
APN = sheet['B3'].value
Alias1_APN = sheet['F3'].value
Alias2_APN = sheet['G3'].value
APNID = sheet['B4'].value
# VRF = sheet['B5'].value
VRF = APN.split('.')[0].upper()
tipo = sheet['B6'].value
GPBSA_peer = sheet['B7'].value
GPBHE_peer = sheet['C7'].value
GPBSA_peer_bkp = sheet['B8'].value
GPBHE_peer_bkp = sheet['C8'].value
GPBSA_range = sheet['B9'].value
GPBHE_range = sheet['C9'].value
GPBSA_RAD1 = sheet['B10'].value
GPBHE_RAD1 = sheet['C10'].value
GPBSA_RAD2 = sheet['B11'].value
GPBHE_RAD2 = sheet['C11'].value
RAD_backup_mode = sheet['F11'].value
RAD_Auth = sheet['B12'].value
RAD_Acct = sheet['B13'].value
RAD_NAS_secret = sheet['B14'].value
auth_type = sheet['B15'].value
Addr_aloc = sheet['B16'].value
policy = sheet['B17'].value
IP_tunnel_BSA = sheet['F17'].value
IP_tunnel_BHE = sheet['F18'].value
PSK = sheet['B18'].value
DNS1 = sheet['F5'].value
DNS2 = sheet['F6'].value
###########################
# redes destino
rede_cli = []
a = 20
# b = 0
while sheet['A' + str(a)].value == None:
    if sheet['B' + str(a - 1)].value != None:
        rede_cli.append(sheet['B' + str(a - 1)].value)
        # b +=1
    a += 1
###########################
# redes destino
URLs_cli = []
a = 22
b = 0
while sheet['A' + str(a)].value == None:
    if sheet['E' + str(a - 1)].value != None:
        URLs_cli.append(sheet['E' + str(a - 1)].value)
        # b +=1
    a += 1
###########################




##################################################################################
##################################################################################



# Inicio

Log = u""
data = datetime.datetime.fromtimestamp(time.time()).strftime('%d-%m-%Y')
# VRF = APN.split('.')[0].upper()

print "\r\n\r\n"
##### Criar Descrição
Log += "Network Engineering - Packet Switching\r\n"
Log += "Projeto:\tProjeto w-VPN " + cliente + " - APN " + APN + "\r\n"
Log += "Data:\t" + data + "\r\n"
Log += "===========================================================================================\r\n\r\n"
Log += u"Descrição\r\n"
Log += "-------------------------------------------------------------------------------------------\r\n"
Log += u"Criação de nova w-VPN\r\n"
Log += "Cliente:\t" + cliente + "\r\n"
Log += "APN:\t\t" + APN + "\r\n"
Log += "APNid:\t\t" + APNID + "\r\n"
Log += "-------------------------------------------------------------------------------------------\r\n\r\n\r\n"

##### Configuração PGWs
if tipo == 'IPsec':
    Log += criar_VRF()
    Log += criar_ACL()
    Log += criar_IKE_peer()
    Log += criar_IPSEC_policy()
    Log += criar_Interface_tunnel_IPSEC()
    Log += criar_IP_pool()
    Log += criar_Route()
    if GPBSA_RAD1 != None:
        Log += criar_GiIf()
        Log += criar_Radius_server_group()
    Log += criar_APN()
    if Alias1_APN != None: Log += criar_Alias_APN()
    Log += "-------------------------------------------------------------------------------------------\r\n\r\n\r\n"
else:
    if tipo == 'GRE':
        Log += criar_VRF()
        Log += criar_Interface_tunnel_GRE()
        Log += criar_IP_pool()
        Log += criar_Route()
        if GPBSA_RAD1 != None:
            Log += criar_GiIf()
            Log += criar_Radius_server_group()
        Log += criar_APN()
        if Alias1_APN != None: Log += criar_Alias_APN()
        Log += "-------------------------------------------------------------------------------------------\r\n\r\n\r\n"
    else:
        if tipo == 'Sem Tunel':
            Log += criar_Filtro()
            if GPBSA_RAD1 != None:
                Log += criar_GiIf()
                Log += criar_Radius_server_group()
            Log += criar_APN()
            if Alias1_APN != None: Log += criar_Alias_APN()
            Log += "-------------------------------------------------------------------------------------------\r\n\r\n\r\n"
        else:
            Log += u"\r\n\r\n***** Configuraçao PGW em desenvolvimento *****\r\n\r\n"

##### Configuração DNS
Log += u"F5 (iDNS & eDNS)\r\n"
Log += "-------------------------------------------------------------------------------------------\r\n"
Log += u"# Configuração DNS\r\n"
Log += u"\r\n# Salvar o script abaixo em um arquivo TXT é utilzar o APP (CVNA_F5_APP_v4.exe) para aplicar a aconfiguração:\r\n"
PGWs = []
if tipo != 'LD':
    PGWs = ['GPBSA1', 'GPBHE1']
else:
    PGWs = ['GPRJO4', 'GPSNE1']
Log += criar_DNS()
Log += "-------------------------------------------------------------------------------------------\r\n\r\n\r\n"

print Log

# --------------------------------
import codecs

arquivo = "./conf_wVPN_" + VRF + ".txt"
# f = codecs.open('./saida_teste.txt','w','utf8')
f = codecs.open(arquivo, 'w', 'utf8')
f.write(Log)  # Stored on disk as UTF-8
f.close()
# --------------------------------
# --------------------------------
print
print 'Arquivo salvo'
print
confirma = raw_input("Tecle algo para terminar...")
print "Concluido..."
# --------------------------------


# =====================================================================--------------------------------------------------
