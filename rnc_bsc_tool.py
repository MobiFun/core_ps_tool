# coding=utf-8
from __future__ import division
import warnings
import openpyxl as pyxl
import openpyxl.styles as pyxlst
from support_functions import find_between, bydigit, byord, remove_duplicate_dicts
import re
import os
import datetime
import exceptions
import json

from support_variables import add_m3de, add_m3lks, add_m3lnk, add_m3rt, add_rnc, add_sccpdpc, add_sccpssn, \
    add_iupaging, slsm_map, add_rac_cname, add_nri_cname, view_dns, rmv_rac_cname, rmv_nri_cname, add_rnc_r7, \
    add_nse, rmv_nse, add_gbiplocendpt, rmv_gbiplocendpt, mnc_select, srvcc_msc

warnings.filterwarnings("ignore")


def main():
    print(u'Welcome to the BSC and RNC creation Tool!')
    user_name = raw_input(u'Please, type in your full name: ')
    while True:
        print(u'{}, choose one of the options below to start: '.format(user_name.split(' ')[0]))
        print(u'1: Create RNCs projects')
        print(u'2: Create BSCs projects')
        print(u'3: Exit')
        choice = raw_input('> ').strip()
        if choice == '1':
            file_input = raw_input(
                u'Type in the path to the Gb Iu Planning files from which you want to create the '
                u'RNCs <separate them with a comma [","]>: ').strip().split(',')

            rnc_name = raw_input(u'Type in the names of the RNCs you want to create a project for '
                                 u'<separate them with a comma [","]> <leave blank to create for all>:').strip()

            expansion = raw_input(u'Is this a new LAC request? <y/n>: ').strip().lower()
            # Check if expansion is valid and rewrite value to use in fulfill_dns_info function
            if expansion == 'y':
                expansion = False
            elif expansion == 'n':
                expansion = True
            else:
                print(u"Sorry, we don't understand ['{}'], please try again.".format(expansion))
                continue
            # Check if rnc_name has any value, if it has, split those values in a list
            if rnc_name:
                rnc_name = rnc_name.split(',')

            file_input_count = 0  # Count number of file inputs
            file_input_dict = dict()  # Create a dict of the file inputs provided by the user
            sgsn_count = 0  # Count number of sgsns
            sgsn_names = dict()  # Create a dict with the names of sgsns to create a folder

            # Populate file_input_dict with the files names
            for files in file_input:
                file_input_dict[file_input_count] = files
                file_input_count += 1

            # Populate sgsn_names dict with SGSN Names
            for key in file_input_dict:
                all_rnc_data = extract_rnc_info_for_project(file_input_dict.get(key), rnc_name=rnc_name)
                for rnc_info in all_rnc_data:
                    sgsn_names[sgsn_count] = rnc_info.get('SGSN Name')
                sgsn_count += 1

            # Auxiliary string to create folder of projects
            sgsn_string = ''
            # Populate the string with SGSNs Names
            for key in sgsn_names:
                sgsn_string += sgsn_names.get(key) + '_'

            sgsn_string = sgsn_string[:-1]  # Remove last character ['_']

            # Create the destination folder using the auxiliary string
            destination_folder = create_projects_destination_folder('RNC', sgsn_string)

            # Start to create RNC projects
            print('')
            print(u'Projects for USNs {} are being created...'.format(sgsn_string))
            print('')
            all_empty_row_data = []
            sgsn_rnc_count = 1
            for key in file_input_dict:
                all_rnc_data = extract_rnc_info_for_project(file_input_dict.get(key), rnc_name=rnc_name)
                if sgsn_rnc_count == 1:
                    for rnc_info in all_rnc_data:
                        dns_sheet_empty_row = 26
                        fallback_dns_sheet_empty_row = 24
                        fallback_empty_row = 15
                        workbook_destination_path = create_rnc_workbook(rnc_info, destination_folder)
                        if workbook_destination_path:
                            srvcc_entries = srvcc_dns_entries_f5(ne_info=rnc_info, expansion=expansion)
                            export_f5_entries(ne_info=rnc_info, full_path=destination_folder,
                                              entries_list=srvcc_entries)
                            fallback_empty_row, dns_sheet_empty_row, fallback_dns_sheet_empty_row = create_rnc_project(
                                workbook_destination_path=workbook_destination_path,
                                user_name=user_name,
                                rnc_data=rnc_info,
                                file_input=file_input_dict.get(key),
                                fallback_empty_row=fallback_empty_row,
                                dns_sheet_empty_row=dns_sheet_empty_row,
                                fallback_dns_sheet_empty_row=fallback_dns_sheet_empty_row,
                                expansion=expansion,
                                sgsn_rnc_count=sgsn_rnc_count)

                            empty_row_dict = {'RNC Name': rnc_info.get('RNC Name'),
                                              'rows': [fallback_empty_row, dns_sheet_empty_row,
                                                       fallback_dns_sheet_empty_row]}

                            all_empty_row_data.append(empty_row_dict)
                        else:
                            raw_input()
                            break
                else:
                    # noinspection PyUnboundLocalVariable
                    for rnc_info in all_rnc_data:
                        all_empty_row_data_refactored = remove_duplicate_dicts(all_empty_row_data)
                        for dicts in all_empty_row_data_refactored:
                            if dicts.get('RNC Name') == rnc_info.get('RNC Name'):
                                workbook_destination_path = create_rnc_workbook(rnc_info, destination_folder)
                                if workbook_destination_path:
                                    fallback_empty_row, dns_sheet_empty_row, fallback_dns_sheet_empty_row = \
                                        create_rnc_project(
                                            workbook_destination_path=workbook_destination_path,
                                            user_name=user_name,
                                            rnc_data=rnc_info,
                                            file_input=file_input_dict.get(key),
                                            fallback_empty_row=dicts.get('rows')[0],
                                            dns_sheet_empty_row=dicts.get('rows')[1],
                                            fallback_dns_sheet_empty_row=dicts.get('rows')[2],
                                            expansion=True,
                                            sgsn_rnc_count=sgsn_rnc_count)

                                    empty_row_dict = {'RNC Name': rnc_info.get('RNC Name'),
                                                      'rows': [fallback_empty_row, dns_sheet_empty_row,
                                                               fallback_dns_sheet_empty_row]}
                                    # noinspection PyUnboundLocalVariable
                                    if sgsn_rnc_count == len(file_input):
                                        all_empty_row_data = [d for d in all_empty_row_data if
                                                              not d.get('RNC Name') == rnc_info.get('RNC Name')]
                                    else:
                                        all_empty_row_data.append(empty_row_dict)
                                        all_empty_row_data = remove_duplicate_dicts(all_empty_row_data)
                                else:
                                    raw_input()
                                    break
                sgsn_rnc_count += 1
            print('')
            print(u'*******************')
            print(u'*Projects Created!*')
            print(u'*******************')
            print('')

        elif choice == '2':
            file_input = raw_input(
                u'Type in the path to the Gb Iu Planning files from which you want to create the '
                u'BSCs <separate them with a comma [","]>: ').strip().split(',')

            bsc_name = raw_input(u'Type in the names of the BSCs you want to create a project for '
                                 u'<separate them with a comma [","]> <leave blank to create for all>:').strip()

            expansion = raw_input(u'Is this a new LAC request? <y/n>: ').strip().lower()

            # Check if expansion is valid and rewrite value to use in fulfill_dns_info function
            if expansion == 'y':
                expansion = False
            elif expansion == 'n':
                expansion = True
            else:
                print(u"Sorry, we don't understand ['{}'], please try again.".format(expansion))
                continue
            # Check if rnc_name has any value, if it has, split those values in a list
            if bsc_name:
                bsc_name = bsc_name.split(',')

            file_input_count = 0  # Count number of file inputs
            file_input_dict = dict()  # Create a dict of the file inputs provided by the user
            sgsn_count = 0  # Count number of sgsns
            sgsn_names = dict()  # Create a dict with the names of sgsns to create a folder

            # Populate file_input_dict with the files names
            for files in file_input:
                file_input_dict[file_input_count] = files
                file_input_count += 1

            # Populate sgsn_names dict with SGSN Names
            for key in file_input_dict:
                all_bsc_data = extract_bsc_info_for_project(file_input_dict.get(key), bsc_name=bsc_name)
                for bsc_info in all_bsc_data:
                    sgsn_names[sgsn_count] = bsc_info.get('SGSN Name')
                sgsn_count += 1

            # Auxiliary string to create folder of projects
            sgsn_string = ''
            # Populate the string with SGSNs Names
            for key in sgsn_names:
                sgsn_string += sgsn_names.get(key) + '_'

            sgsn_string = sgsn_string[:-1]  # Remove last character ['_']

            # Create the destination folder using the auxiliary string
            destination_folder = create_projects_destination_folder('BSC', sgsn_string)

            # Start to create BSC projects
            print('')
            print(u'Projects for USNs {} are being created...'.format(sgsn_string))
            print('')

            all_empty_row_data = []
            sgsn_bsc_count = 1
            for key in file_input_dict:
                all_bsc_data = extract_bsc_info_for_project(file_input_dict.get(key), bsc_name=bsc_name)
                if sgsn_bsc_count == 1:
                    for bsc_info in all_bsc_data:
                        dns_sheet_empty_row = 26
                        fallback_dns_sheet_empty_row = 24
                        fallback_empty_row = 15
                        workbook_destination_path = create_bsc_workbook(bsc_info, destination_folder)
                        if workbook_destination_path:
                            srvcc_entries = srvcc_dns_entries_f5(ne_info=bsc_info, expansion=expansion)
                            export_f5_entries(ne_info=bsc_info, full_path=destination_folder,
                                              entries_list=srvcc_entries)
                            fallback_empty_row, dns_sheet_empty_row, fallback_dns_sheet_empty_row = create_bsc_project(
                                workbook_destination_path=workbook_destination_path,
                                user_name=user_name,
                                bsc_data=bsc_info,
                                file_input=file_input_dict.get(key),
                                fallback_empty_row=fallback_empty_row,
                                dns_sheet_empty_row=dns_sheet_empty_row,
                                fallback_dns_sheet_empty_row=fallback_dns_sheet_empty_row,
                                expansion=expansion,
                                sgsn_bsc_count=sgsn_bsc_count)

                            empty_row_dict = {'BSC Name': bsc_info.get('BSC Name'),
                                              'rows': [fallback_empty_row, dns_sheet_empty_row,
                                                       fallback_dns_sheet_empty_row]}

                            all_empty_row_data.append(empty_row_dict)
                        else:
                            raw_input('Press Enter! 1')
                            break
                else:
                    for bsc_info in all_bsc_data:
                        all_empty_row_data_refactored = remove_duplicate_dicts(all_empty_row_data)
                        for dicts in all_empty_row_data_refactored:
                            if dicts.get('BSC Name') == bsc_info.get('BSC Name'):
                                workbook_destination_path = create_bsc_workbook(bsc_info, destination_folder)
                                if workbook_destination_path:
                                    fallback_empty_row, dns_sheet_empty_row, fallback_dns_sheet_empty_row = \
                                        create_bsc_project(
                                            workbook_destination_path=workbook_destination_path,
                                            user_name=user_name,
                                            bsc_data=bsc_info,
                                            file_input=file_input_dict.get(key),
                                            fallback_empty_row=dicts.get('rows')[0],
                                            dns_sheet_empty_row=dicts.get('rows')[1],
                                            fallback_dns_sheet_empty_row=dicts.get('rows')[2],
                                            expansion=True,
                                            sgsn_bsc_count=sgsn_bsc_count)

                                    empty_row_dict = {'BSC Name': bsc_info.get('BSC Name'),
                                                      'rows': [fallback_empty_row, dns_sheet_empty_row,
                                                               fallback_dns_sheet_empty_row]}
                                    if sgsn_bsc_count == len(file_input):
                                        all_empty_row_data = [d for d in all_empty_row_data if
                                                              not d.get('BSC Name') == bsc_info.get('BSC Name')]
                                    else:
                                        all_empty_row_data.append(empty_row_dict)
                                        all_empty_row_data = remove_duplicate_dicts(all_empty_row_data)
                                else:
                                    raw_input('Press Enter!')
                                    break
                sgsn_bsc_count += 1

            print('')
            print(u'*******************')
            print(u'*Projects Created!*')
            print(u'*******************')
            print('')

        elif choice == '3':
            print(u'Thank you for using the RNC and BSC creation Tool!')
            print(u'If you have any suggestions please send an e-mail to: decastromonteiro@gmail.com')
            raw_input()
            break

        else:
            print(u"Sorry, we don't understand ['{}'], please try again.".format(choice))
            continue


def extract_rnc_info_for_project(file_input, rnc_name):
    digit_pattern = "[-+]?\d+[\.]?\d*"
    active_workbook = pyxl.load_workbook(file_input, data_only=False)
    iuc = active_workbook.get_sheet_by_name('Iu_C')
    merged = iuc.merged_cell_ranges
    merged = sorted(merged, key=lambda x: byord(x))
    merged = sorted(merged, key=lambda x: bydigit(x))
    counter = 0

    all_rnc_data = list()

    cell_header = {
        0: 'RNC Name',
        1: 'Vendor',
        2: 'RNCX',
        3: 'RNCID',
        4: 'DPC',
        5: 'DPX',
        6: 'RNC MCC',
        7: 'RNC MNC',
        8: 'NI',
        9: 'Iu-Flex',
        10: 'Ran Sharing',
        11: 'DT',
        12: 'Support R7 QoS',
        13: 'RAB QOS',
        14: 'RNC Version',
        15: 'SGSN Name',
        16: 'NRI',
        17: 'OPC',
        18: 'OPX',
        19: 'NI Logical Link',
        20: 'LEX',
        21: 'DEX',
        22: 'LSX',
        23: 'DET',
        24: 'SSNX (RANAP)',
        25: 'SSNX (SCMG)',
        26: 'SLSM'
    }
    rnc_info = dict()
    if not rnc_name:
        for cell in merged[2:]:

            cell_first_index = find_between(cell, '', ':')
            int_cell_first_index = int(re.findall(digit_pattern, cell_first_index)[0])
            cell_last_index = find_between(cell, ':', '')
            int_cell_last_index = int(re.findall(digit_pattern, cell_last_index)[0])

            if cell_header.get(counter % 27) == 'RNC Name':
                rnc_info[cell_header.get(counter % 27)] = str(iuc[cell_first_index].value)
                rnc_info['Coordinate'] = (str(int_cell_first_index), str(int_cell_last_index))
                lai_list = list()
                for lai_count in xrange(int_cell_first_index, int_cell_last_index + 1):
                    lac_cell = str(iuc['K{}'.format(lai_count)].value)
                    rac_cell = str(iuc['L{}'.format(lai_count)].value) if iuc[
                        'L{}'.format(lai_count)].value else '0'
                    if lac_cell != 'None':
                        lai_list.append({'LAC': lac_cell, 'RAC': rac_cell})

                rnc_info['LAI'] = lai_list
                m3lnk_list = list()
                for m3lnk_count in xrange(int_cell_first_index, int_cell_last_index + 1):
                    srn = str(iuc['W{}'.format(m3lnk_count)].value)
                    sn = str(iuc['X{}'.format(m3lnk_count)].value)
                    lnk = str(iuc['Y{}'.format(m3lnk_count)].value)
                    sgsn_ip1 = str(iuc['AG{}'.format(m3lnk_count)].value)
                    sgsn_ip2 = str(iuc['AH{}'.format(m3lnk_count)].value)
                    sgsn_port = str(iuc['AI{}'.format(m3lnk_count)].value)
                    rnc_ip1 = str(iuc['AJ{}'.format(m3lnk_count)].value)
                    rnc_ip2 = str(iuc['AK{}'.format(m3lnk_count)].value)
                    rnc_port = str(iuc['AL{}'.format(m3lnk_count)].value)
                    if sgsn_ip1 != 'None':
                        m3lnk_list.append(
                            {
                                'SRN': srn, 'SN': sn, 'LNK': lnk, 'LOCIPV41': sgsn_ip1, 'LOCIPV42': sgsn_ip2,
                                'LOCPORT': sgsn_port, 'PEERIPV41': rnc_ip1, 'PEERIPV42': rnc_ip2, 'PEERPORT': rnc_port
                            }
                        )

                rnc_info['M3LNK'] = m3lnk_list
            else:
                rnc_info[cell_header.get(counter % 27)] = str(iuc[cell_first_index].value)

            if counter != 0:
                if counter % 27 == 26:
                    all_rnc_data.append(rnc_info)
                    rnc_info = dict()
                    counter += 1
                else:
                    counter += 1
            else:
                counter += 1
    else:
        for rnc in rnc_name:
            for cell in merged[2:]:

                cell_first_index = find_between(cell, '', ':')
                int_cell_first_index = int(re.findall(digit_pattern, cell_first_index)[0])
                cell_last_index = find_between(cell, ':', '')
                int_cell_last_index = int(re.findall(digit_pattern, cell_last_index)[0])

                if cell_header.get(counter % 27) == 'RNC Name':
                    rnc_info[cell_header.get(counter % 27)] = str(iuc[cell_first_index].value)
                    rnc_info['Coordinate'] = (str(int_cell_first_index), str(int_cell_last_index))
                    lai_list = list()
                    for lai_count in xrange(int_cell_first_index, int_cell_last_index + 1):
                        lac_cell = str(iuc['K{}'.format(lai_count)].value)
                        rac_cell = str(iuc['L{}'.format(lai_count)].value) if iuc[
                            'L{}'.format(lai_count)].value else '0'
                        if lac_cell != 'None':
                            lai_list.append({'LAC': lac_cell, 'RAC': rac_cell})

                    rnc_info['LAI'] = lai_list
                    m3lnk_list = list()
                    for m3lnk_count in xrange(int_cell_first_index, int_cell_last_index + 1):
                        srn = str(iuc['W{}'.format(m3lnk_count)].value)
                        sn = str(iuc['X{}'.format(m3lnk_count)].value)
                        lnk = str(iuc['Y{}'.format(m3lnk_count)].value)
                        sgsn_ip1 = str(iuc['AG{}'.format(m3lnk_count)].value)
                        sgsn_ip2 = str(iuc['AH{}'.format(m3lnk_count)].value)
                        sgsn_port = str(iuc['AI{}'.format(m3lnk_count)].value)
                        rnc_ip1 = str(iuc['AJ{}'.format(m3lnk_count)].value)
                        rnc_ip2 = str(iuc['AK{}'.format(m3lnk_count)].value)
                        rnc_port = str(iuc['AL{}'.format(m3lnk_count)].value)
                        if sgsn_ip1 != 'None':
                            m3lnk_list.append(
                                {
                                    'SRN': srn, 'SN': sn, 'LNK': lnk, 'LOCIPV41': sgsn_ip1, 'LOCIPV42': sgsn_ip2,
                                    'LOCPORT': sgsn_port, 'PEERIPV41': rnc_ip1, 'PEERIPV42': rnc_ip2,
                                    'PEERPORT': rnc_port
                                }
                            )

                    rnc_info['M3LNK'] = m3lnk_list
                else:
                    rnc_info[cell_header.get(counter % 27)] = str(iuc[cell_first_index].value)

                if counter != 0:
                    if counter % 27 == 26:
                        if rnc_info.get('RNC Name') == rnc:
                            all_rnc_data.append(rnc_info)
                        rnc_info = dict()
                        counter += 1
                    else:
                        counter += 1
                else:
                    counter += 1

    return all_rnc_data


def extract_bsc_info_for_project(gb_iu_planning_input, bsc_name):
    digit_pattern = '[-+]?\d+[\.]?\d*'
    active_workbook = pyxl.load_workbook(gb_iu_planning_input, data_only=True)
    gbo_ip = active_workbook.get_sheet_by_name('GBoIP')
    merged = gbo_ip.merged_cell_ranges
    merged = sorted(merged, key=lambda x: byord(x))
    merged = sorted(merged, key=lambda x: bydigit(x))
    counter = 0

    all_bsc_data = list()

    cell_header = {
        0: 'BSC Name',
        1: 'TX',
        2: 'NSE Conf Type',
        3: 'SGSN Name',
        4: 'NRI',
        5: 'GB Flex'
    }
    bsc_info = dict()
    if not bsc_name:
        for cell in merged[18:]:
            cell_first_index = find_between(cell, '', ':')
            int_cell_first_index = int(re.findall(digit_pattern, cell_first_index)[0])
            cell_last_index = find_between(cell, ':', '')
            int_cell_last_index = int(re.findall(digit_pattern, cell_last_index)[0])
            bsc_info[cell_header.get(counter % 6)] = str(gbo_ip[cell_first_index].value)

            if cell_header.get(counter % 6) == 'BSC Name':
                bsc_info['Coordinate'] = (str(int_cell_first_index), str(int_cell_last_index))
                lai_list = list()
                for lai_count in xrange(int_cell_first_index, int_cell_last_index + 1):
                    lac_cell = str(gbo_ip['E{}'.format(lai_count)].value)
                    rac_cell = str(gbo_ip['F{}'.format(lai_count)].value) if gbo_ip[
                        'F{}'.format(lai_count)].value else '0'
                    if lac_cell != 'None':
                        lai_list.append({'LAC': lac_cell, 'RAC': rac_cell})

                bsc_info['LAI'] = lai_list
                nsei_list = list()
                for nsei_count in xrange(int_cell_first_index, int_cell_last_index + 1):
                    nsei = str(gbo_ip['C{}'.format(nsei_count)].value)
                    if nsei != 'None':
                        nsei_list.append(nsei)

                bsc_info['NSEI'] = nsei_list
                gblocalend_list = list()
                for gblocalend_count in xrange(int_cell_first_index, int_cell_last_index + 1):
                    srn = str(gbo_ip['L{}'.format(gblocalend_count)].value)
                    sn = str(gbo_ip['M{}'.format(gblocalend_count)].value)
                    sgsn_ip1 = str(gbo_ip['N{}'.format(gblocalend_count)].value)
                    sgsn_ip2 = str(gbo_ip['O{}'.format(gblocalend_count)].value)
                    sgsn_port = str(gbo_ip['P{}'.format(gblocalend_count)].value)
                    if sgsn_ip1 != 'None':
                        gblocalend_list.append(
                            {'SRN': srn, 'SN': sn, 'LOCALIP1': sgsn_ip1, 'LOCALIP2': sgsn_ip2, 'LOCALPORT': sgsn_port}
                        )
                bsc_info['GBLOCALENDPOINT'] = gblocalend_list
            else:
                bsc_info[cell_header.get(counter % 6)] = str(gbo_ip[cell_first_index].value)

            if counter != 0:
                if counter % 6 == 5:
                    all_bsc_data.append(bsc_info)
                    bsc_info = dict()
                    counter += 1
                else:
                    counter += 1
            else:
                counter += 1
    else:
        for bsc in bsc_name:
            for cell in merged[18:]:
                cell_first_index = find_between(cell, '', ':')
                int_cell_first_index = int(re.findall(digit_pattern, cell_first_index)[0])
                cell_last_index = find_between(cell, ':', '')
                int_cell_last_index = int(re.findall(digit_pattern, cell_last_index)[0])
                bsc_info[cell_header.get(counter % 6)] = str(gbo_ip[cell_first_index].value)

                if cell_header.get(counter % 6) == 'BSC Name':
                    bsc_info['Coordinate'] = (str(int_cell_first_index), str(int_cell_last_index))
                    lai_list = list()
                    for lai_count in xrange(int_cell_first_index, int_cell_last_index + 1):
                        lac_cell = str(gbo_ip['E{}'.format(lai_count)].value)
                        rac_cell = str(gbo_ip['F{}'.format(lai_count)].value) if gbo_ip[
                            'F{}'.format(lai_count)].value else '0'
                        if lac_cell != 'None':
                            lai_list.append({'LAC': lac_cell, 'RAC': rac_cell})

                    bsc_info['LAI'] = lai_list
                    nsei_list = list()
                    for nsei_count in xrange(int_cell_first_index, int_cell_last_index + 1):
                        nsei = str(gbo_ip['C{}'.format(nsei_count)].value)
                        if nsei != 'None':
                            nsei_list.append(nsei)

                    bsc_info['NSEI'] = nsei_list
                    gblocalend_list = list()
                    for gblocalend_count in xrange(int_cell_first_index, int_cell_last_index + 1):
                        srn = str(gbo_ip['L{}'.format(gblocalend_count)].value)
                        sn = str(gbo_ip['M{}'.format(gblocalend_count)].value)
                        sgsn_ip1 = str(gbo_ip['N{}'.format(gblocalend_count)].value)
                        sgsn_ip2 = str(gbo_ip['O{}'.format(gblocalend_count)].value)
                        sgsn_port = str(gbo_ip['P{}'.format(gblocalend_count)].value)
                        if sgsn_ip1 != 'None':
                            gblocalend_list.append(
                                {'SRN': srn, 'SN': sn, 'LOCALIP1': sgsn_ip1, 'LOCALIP2': sgsn_ip2,
                                 'LOCALPORT': sgsn_port}
                            )
                    bsc_info['GBLOCALENDPOINT'] = gblocalend_list
                else:
                    bsc_info[cell_header.get(counter % 6)] = str(gbo_ip[cell_first_index].value)

                if counter != 0:
                    if counter % 6 == 5:
                        if bsc_info.get('BSC Name') == bsc:
                            all_bsc_data.append(bsc_info)
                        bsc_info = dict()
                        counter += 1
                    else:
                        counter += 1
                else:
                    counter += 1

    return all_bsc_data


def create_rnc_project(workbook_destination_path, user_name, rnc_data, file_input, fallback_empty_row,
                       dns_sheet_empty_row, fallback_dns_sheet_empty_row, expansion, sgsn_rnc_count):
    digit_pattern = "[-+]?\d+[\.]?\d*"  # Regex para encontrar digitos numa string
    wb_from = pyxl.load_workbook(file_input, keep_vba=True)  # Abrir o Gb_Iu_Planning mantendo o codigo VBA
    ws_from = wb_from.get_sheet_by_name('Iu_C')  # Selecionar a Aba "Iu_C"
    # Abrir o workbook da RNC
    workbook = pyxl.load_workbook(workbook_destination_path)
    # Criar aba com nome do SGSN
    ws_to = workbook.create_sheet(title='Projeto {}'.format(rnc_data.get('SGSN Name')), index=2)
    # ws_to.title = 'Projeto {}'.format(rnc_data.get('SGSN Name'))  # Criar aba com nome do SGSN
    ws_to_fallback = workbook.get_sheet_by_name('Fallback')  # Selecionar aba Fallback
    merged = ws_from.merged_cell_ranges  # Encontrar todas as Células Mescladas
    merged = sorted(merged, key=lambda x: byord(x))  # Ordernar a lista com base nas colunas
    merged = sorted(merged, key=lambda x: bydigit(x))  # Ordernar a lista com base nas linhas

    # Copiar os cabeçalhos do Gb_Iu_Planning para a nova aba do novo Workbook
    for rrow in ws_from.iter_rows('A1:AQ2'):
        for cell in rrow:
            ws_to[cell.coordinate].value = cell.value
            ws_to[cell.coordinate].style = cell.style

    # Mesclar as células do cabeçalho
    for merged_cells in merged[:2]:
        ws_to.merge_cells(merged_cells)

    # Copiar as células com informações da RNC alvo para a nova aba do novo Workbook
    row_count = 0
    for rrow in ws_from.iter_rows(
            'A{}:AQ{}'.format(rnc_data.get('Coordinate')[0], rnc_data.get('Coordinate')[1])):
        for cell in rrow:
            ranges = re.findall(digit_pattern, cell.coordinate)  # Encontrar as linhas correspondentes das celulas

            coordinate_normalized = 3 + row_count

            # Criar a nova coordenada já normalizada
            cell_coordinate = cell.coordinate.replace(ranges[0], str(coordinate_normalized))
            # Copiar os valores das células
            ws_to[cell_coordinate] = cell.value
            # Copiar os estilos das células
            ws_to[cell_coordinate].style = cell.style
        row_count += 1

    # Encontrar o INDEX da célula RNC Name da RNC alvo
    start_row = rnc_data.get('Coordinate')[0]
    end_row = rnc_data.get('Coordinate')[1]
    rnc_name_index = merged.index('A{}:A{}'.format(start_row, end_row))

    # Mesclar as células que contém informação da RNC alvo

    for merged_cells in merged[rnc_name_index:rnc_name_index + 27]:
        ranges = re.findall(digit_pattern, merged_cells)  # Encontrar as linhas correspondentes das celulas
        # Normalizar a coordenada para iniciar a Mescla a partir da célula A3
        first_range_normalized = 3
        # Normalizar a coordenada para finalizar a Mescla na célula correta
        second_range_normalized = first_range_normalized + (int(ranges[1]) - int(ranges[0]))

        temp_coordinate = merged_cells.replace(ranges[0], str(first_range_normalized))
        final_coordinate = temp_coordinate.replace(ranges[1], str(second_range_normalized))
        # Mesclar as células com as coordenadas normalizadas
        ws_to.merge_cells(final_coordinate)

    fulfill_cover_info(workbook, 'Projeto Integração RNC - {}'.format(rnc_data.get('RNC Name')), user_name)
    fallback_empty_row = fulfill_rnc_info(ws_to, ws_to_fallback, rnc_data, fallback_empty_row)
    if sgsn_rnc_count == 1:
        dns_sheet_empty_row, fallback_dns_sheet_empty_row = fulfill_dns_info(workbook, rnc_data,
                                                                             dns_sheet_empty_row,
                                                                             fallback_dns_sheet_empty_row,
                                                                             expansion)
    # Salvar o novo Workbook
    workbook.save(filename=workbook_destination_path)

    return fallback_empty_row, dns_sheet_empty_row, fallback_dns_sheet_empty_row


def create_rnc_workbook(rnc_data, full_path):
    date = datetime.datetime.now()
    destination_path = os.path.join(full_path,
                                    'Projeto Integracao RNC_{}_{}.xlsx'.format(rnc_data.get('RNC Name'),
                                                                               date.strftime('%d%m%Y'))
                                    )
    if not os.path.exists(destination_path):
        base_folder = os.path.abspath(os.getcwd())
        excel_template = os.path.join(base_folder, 'excel_templates\Project_Template.xlsx')
        try:
            wb = pyxl.load_workbook(filename=excel_template)
            wb.save(filename=destination_path)
            return destination_path
        except exceptions.IOError as error:
            if error[0] == 2:
                print('Sorry, but the path to excel_templates does not exist.')
                print('Path expected: "{}"'.format(excel_template))
            else:
                print('Sorry, there was an IOError.')
                print('ERROR: {}'.format(error))
    else:
        return destination_path


def create_projects_destination_folder(element_type, sgsn_string):
    base_folder = os.path.abspath(os.getcwd())
    destination_folder = 'Projetos {} - {}'.format(element_type, sgsn_string)
    full_path = os.path.join(base_folder, destination_folder)
    if not os.path.exists(full_path):
        os.makedirs(full_path)
    return full_path


def fulfill_cover_info(active_workbook, project_title, user_name):
    """This function receives a full path to a Workbook, the Project Title and
    the User Name to fulfill the Cover of the Excel Workbook"""
    date = datetime.datetime.now()  # Today date information
    # Cover Sheet
    cover = active_workbook.get_sheet_by_name('Capa')
    cover['F9'] = date.strftime('%d/%m/%Y')  # Escrever a data na célula F9
    cover['J6'] = project_title  # Escrever o Título do projeto na célula J6
    cover['H13'] = user_name.strip()  # Escrever o nome do usuário na célula H13
    cover['D21'] = user_name.strip()  # Escrever o nome do usuário na célula D21
    cover['H21'] = date.strftime('%d/%m/%Y')  # Escrever a data na célula H21
    cover['K21'] = '1'  # Escrever o número da versão na cpelula K21
    cover['M21'] = 'Primeira versão'  # Escrever observação na célula M21


def fulfill_rnc_info(rnc_worksheet, fallback, rnc_info, fallback_empty_row):
    # Define font style
    font = pyxlst.Font(name='Arial', size=11, bold=True, underline="single")

    # RNC Sheet

    empty_row = rnc_worksheet.max_row + 3
    rnc_worksheet['A{}'.format(empty_row)] = '/* To use on BATCH COMMAND */'
    rnc_worksheet['A{}'.format(empty_row)].font = font
    empty_row = rnc_worksheet.max_row + 1
    rnc_worksheet['A{}'.format(empty_row)] = 'USE ME:MEID = 8;'

    # Configure M3UA Destination Entity
    empty_row = rnc_worksheet.max_row + 2
    rnc_worksheet['A{}'.format(empty_row)] = '/* 1 - Add an M3UA destination entity */'
    rnc_worksheet['A{}'.format(empty_row)].font = font
    empty_row = rnc_worksheet.max_row + 1
    rnc_worksheet['A{}'.format(empty_row)] = add_m3de.format(DEX=rnc_info.get('DEX'),
                                                             LEX=rnc_info.get('LEX'),
                                                             DPC=rnc_info.get('DPC'),
                                                             DET=rnc_info.get('DET'),
                                                             RNC_Name=rnc_info.get('RNC Name'))
    # Configure M3UA Signaling Link Sets
    empty_row = rnc_worksheet.max_row + 2
    rnc_worksheet['A{}'.format(empty_row)] = '/* 2 - Add M3UA signaling link sets */'
    rnc_worksheet['A{}'.format(empty_row)].font = font
    empty_row = rnc_worksheet.max_row + 1
    rnc_worksheet['A{}'.format(empty_row)] = add_m3lks.format(LSX=rnc_info.get('LSX'),
                                                              DEX=rnc_info.get('DEX'),
                                                              DET=rnc_info.get('DET'),
                                                              Share_Type='LOADSHARE',
                                                              RNC_Name=rnc_info.get('RNC Name'),
                                                              SLSM=slsm_map.get(int(rnc_info.get('SLSM'))))
    # Configure M3UA Signaling Routes
    empty_row = rnc_worksheet.max_row + 2
    rnc_worksheet['A{}'.format(empty_row)] = '/* 3 - Add M3UA signaling routes */'
    rnc_worksheet['A{}'.format(empty_row)].font = font
    empty_row = rnc_worksheet.max_row + 1
    rnc_worksheet['A{}'.format(empty_row)] = add_m3rt.format(RTX=rnc_info.get('RNCX'),
                                                             DEX=rnc_info.get('DEX'),
                                                             LSX=rnc_info.get('LSX'),
                                                             RNC_Name=rnc_info.get('RNC Name'))
    # Configure M3UA Link
    empty_row = rnc_worksheet.max_row + 2
    rnc_worksheet['A{}'.format(empty_row)] = '/* 4 - Add an M3UA link */'
    rnc_worksheet['A{}'.format(empty_row)].font = font
    empty_row = rnc_worksheet.max_row + 1
    m3lnk_count = len(rnc_info.get('M3LNK'))
    for x in xrange(m3lnk_count):
        rnc_worksheet['A{}'.format(empty_row + x)] = add_m3lnk.format(SRN=rnc_info.get('M3LNK')[x].get('SRN'),
                                                                      SN=rnc_info.get('M3LNK')[x].get('SN'),
                                                                      LNK=rnc_info.get('M3LNK')[x].get('LNK'),
                                                                      LOCIPV41=rnc_info.get('M3LNK')[x].get('LOCIPV41'),
                                                                      LOCIPV42=rnc_info.get('M3LNK')[x].get('LOCIPV42'),
                                                                      LOCPORT=rnc_info.get('M3LNK')[x].get('LOCPORT'),
                                                                      PEERIPV41=rnc_info.get('M3LNK')[x].get(
                                                                          'PEERIPV41'),
                                                                      PEERIPV42=rnc_info.get('M3LNK')[x].get(
                                                                          'PEERIPV42'),
                                                                      PEERPORT=rnc_info.get('M3LNK')[x].get('PEERPORT'),
                                                                      LSX=rnc_info.get('LSX'),
                                                                      RNC_Name=rnc_info.get('RNC Name')
                                                                      )
    # Configure SCCP Layer Data
    empty_row = rnc_worksheet.max_row + 2
    rnc_worksheet['A{}'.format(empty_row)] = '/* 5 - Configure the SCCP layer data */'
    rnc_worksheet['A{}'.format(empty_row)].font = font
    empty_row = rnc_worksheet.max_row + 1
    rnc_worksheet['A{}'.format(empty_row)] = add_sccpdpc.format(DPX=rnc_info.get('RNCX'),
                                                                OPX=rnc_info.get('OPX'),
                                                                DPC=rnc_info.get('DPC'),
                                                                RNC_Name=rnc_info.get('RNC Name'))
    # Configure Subsystem SCCP
    empty_row = rnc_worksheet.max_row + 2
    rnc_worksheet['A{}'.format(empty_row)] = '/* 6 - Configure the SCCP subsystem */'
    rnc_worksheet['A{}'.format(empty_row)].font = font
    empty_row = rnc_worksheet.max_row + 1
    rnc_worksheet['A{}'.format(empty_row)] = add_sccpssn.format(SSNX=rnc_info.get('SSNX (RANAP)'),
                                                                SSN='RANAP',
                                                                DPC=rnc_info.get('DPC'),
                                                                OPC=rnc_info.get('OPC'),
                                                                RNC_Name=rnc_info.get('RNC Name'))

    rnc_worksheet['A{}'.format(empty_row + 1)] = add_sccpssn.format(SSNX=rnc_info.get('SSNX (SCMG)'),
                                                                    SSN='SCMG',
                                                                    DPC=rnc_info.get('DPC'),
                                                                    OPC=rnc_info.get('OPC'),
                                                                    RNC_Name=rnc_info.get('RNC Name'))
    # Configure RNC
    empty_row = rnc_worksheet.max_row + 2
    rnc_worksheet['A{}'.format(empty_row)] = '/* 7 - Configure the RNC information */'
    rnc_worksheet['A{}'.format(empty_row)].font = font
    empty_row = rnc_worksheet.max_row + 1
    if rnc_info.get('RNC Version') == 'R8':
        rnc_worksheet['A{}'.format(empty_row)] = add_rnc.format(RNCX=rnc_info.get('RNCX'),
                                                                RNC_Name=rnc_info.get('RNC Name'),
                                                                RNCMCC=rnc_info.get('RNC MCC'),
                                                                RNCMNC=rnc_info.get('RNC MNC'),
                                                                RNCINDEX=rnc_info.get('RNCID'),
                                                                NI=rnc_info.get('NI'),
                                                                DPC=rnc_info.get('DPC'),
                                                                Iu_Flex=rnc_info.get('Iu-Flex'),
                                                                RanSharing=rnc_info.get('Ran Sharing'),
                                                                DT=rnc_info.get('DT'),
                                                                R7_QOS=rnc_info.get('Support R7 QoS'),
                                                                RAB_QOS=rnc_info.get('RAB QOS')
                                                                )
    else:
        rnc_worksheet['A{}'.format(empty_row)] = add_rnc_r7.format(RNCX=rnc_info.get('RNCX'),
                                                                   RNC_Name=rnc_info.get('RNC Name'),
                                                                   RNCMCC=rnc_info.get('RNC MCC'),
                                                                   RNCMNC=rnc_info.get('RNC MNC'),
                                                                   RNCINDEX=rnc_info.get('RNCID'),
                                                                   NI=rnc_info.get('NI'),
                                                                   DPC=rnc_info.get('DPC'),
                                                                   Iu_Flex=rnc_info.get('Iu-Flex'),
                                                                   RanSharing=rnc_info.get('Ran Sharing'),
                                                                   R7_QOS=rnc_info.get('Support R7 QoS'),
                                                                   RAB_QOS=rnc_info.get('RAB QOS')
                                                                   )

    # Configure IUPaging
    empty_row = rnc_worksheet.max_row + 2
    rnc_worksheet['A{}'.format(empty_row)] = '/* 8 - Configure the 3G paging table */'
    rnc_worksheet['A{}'.format(empty_row)].font = font
    empty_row = rnc_worksheet.max_row + 1
    lai_count = len(rnc_info.get('LAI'))
    plmn = str(rnc_info.get('RNC MCC')) + str(rnc_info.get('RNC MNC'))
    for x in xrange(lai_count):
        rnc_worksheet['A{}'.format(empty_row + x)] = add_iupaging.format(
            LAI=(plmn + format(int(rnc_info.get('LAI')[x].get('LAC')), '04x')).upper(),
            RAC='0x' + format(int(rnc_info.get('LAI')[x].get('RAC')), '02x').upper(),
            RNCX=rnc_info.get('RNCX')
        )

    # Fallback Sheet
    fallback['E{}'.format(fallback_empty_row)] = '/* {} */'.format(rnc_info.get('SGSN Name'))
    fallback['E{}'.format(fallback_empty_row)].font = font
    fallback_empty_row += 1
    for x in xrange(lai_count):
        fallback[
            'E{}'.format(fallback_empty_row)] = 'RMV IUPAGING: LAI="{LAI}", RAC="{RAC}", RNCINDEX={RNCX};'.format(
            LAI=(plmn + format(int(rnc_info.get('LAI')[x].get('LAC')), '04x')).upper(),
            RAC='0x' + format(int(rnc_info.get('LAI')[x].get('RAC')), '02x').upper(),
            RNCX=rnc_info.get('RNCX')
        )
        fallback_empty_row += 1

    fallback['E{}'.format(fallback_empty_row)] = 'RMV RNC: RNCX={RNCX};'.format(RNCX=rnc_info.get('RNCX'))
    fallback_empty_row += 1
    fallback['E{}'.format(fallback_empty_row)] = 'RMV SCCPSSN: SSNX={SSNX};'.format(SSNX=rnc_info.get('SSNX (RANAP)'))
    fallback_empty_row += 1
    fallback['E{}'.format(fallback_empty_row)] = 'RMV SCCPSSN: SSNX={SSNX};'.format(SSNX=rnc_info.get('SSNX (SCMG)'))
    fallback_empty_row += 1
    fallback['E{}'.format(fallback_empty_row)] = 'RMV SCCPDPC: DPX={DPX};'.format(DPX=rnc_info.get('RNCX'))
    fallback_empty_row += 1
    for x in xrange(m3lnk_count):
        fallback['E{}'.format(fallback_empty_row)] = 'RMV M3LNK: SRN={SRN}, SN={SN}, LNK={LNK};'.format(
            SRN=rnc_info.get('M3LNK')[x].get('SRN'), SN=rnc_info.get('M3LNK')[x].get('SN'),
            LNK=rnc_info.get('M3LNK')[x].get('LNK')
        )
        fallback_empty_row += 1

    fallback['E{}'.format(fallback_empty_row)] = 'RMV M3RT: RTX={RTX};'.format(RTX=rnc_info.get('RNCX'))
    fallback_empty_row += 1
    fallback['E{}'.format(fallback_empty_row)] = 'RMV M3LKS: LSX={LSX};'.format(LSX=rnc_info.get('LSX'))
    fallback_empty_row += 1
    fallback['E{}'.format(fallback_empty_row)] = 'RMV M3DE: DEX={DEX};'.format(DEX=rnc_info.get('DEX'))
    fallback_empty_row += 2

    return fallback_empty_row


# noinspection PyUnusedLocal
def fulfill_dns_info(workbook, ne_info, dns_sheet_empty_row, fallback_dns_sheet_empty_row, expansion=True):
    # Define Font
    font = pyxlst.Font(name='Arial', size=11, bold=True, underline="single")
    # Create iDNS Sheet Object
    dns_sheet = workbook.get_sheet_by_name('iDNS')
    # Create Fallback iDNS Sheet Object
    fallback_dns_sheet = workbook.get_sheet_by_name('Fallback iDNS')
    # Define the number of RACxLAC relations to configure
    lai_count = len(ne_info.get('LAI'))
    # Get RNC or BSC Name
    ne_name = ne_info.get('RNC Name')
    project_name = 'Projeto Integracao RNC_{}.txt'.format(ne_name)
    fallback_project_name = 'FALLBACK_Projeto Integracao RNC_{}.txt'.format(ne_name)
    if not ne_name:
        ne_name = ne_info.get('BSC Name')
        project_name = 'Projeto Integracao BSC_{}.txt'.format(ne_name)
        fallback_project_name = 'FALLBACK_Projeto Integracao BSC_{}.txt'.format(ne_name)

    # iDNS Sheet
    dns_sheet['F{}'.format(dns_sheet_empty_row)] = '/* Projeto DNS F5 */'
    dns_sheet['F{}'.format(dns_sheet_empty_row)].font = font
    dns_sheet_empty_row += 2
    dns_sheet['F{}'.format(dns_sheet_empty_row)] = '{}'.format(project_name)
    dns_sheet_empty_row += 1

    """
    Script Deprecreated due to SWAP of Huawei DNS
    for x in xrange(lai_count):
        mnc = mnc_select.get(ne_info.get('LAI')[x].get('LAC')[-2:-1])
        dns_sheet['F{}'.format(dns_sheet_empty_row)] = '/* LAC: {}  RAC: {} */'.format(
            ne_info.get('LAI')[x].get('LAC'),
            ne_info.get('LAI')[x].get('RAC'))
        dns_sheet['F{}'.format(dns_sheet_empty_row)].font = font
        dns_sheet_empty_row += 1
        if not expansion:
            for views in view_dns:
                dns_sheet['F{}'.format(dns_sheet_empty_row)] = add_rac_cname.format(
                    MNC=mnc,
                    view_name=views,
                    RAC=format(int(ne_info.get('LAI')[x].get('RAC')), '02x').upper(),
                    LAC=format(int(ne_info.get('LAI')[x].get('LAC')), '04x').upper(),
                    SGSN=ne_info.get('SGSN Name')
                )
                dns_sheet_empty_row += 1
        dns_sheet_empty_row += 2
        dns_sheet['F{}'.format(dns_sheet_empty_row)] = '/* NRI: {} */'.format(ne_info.get('NRI'))
        dns_sheet['F{}'.format(dns_sheet_empty_row)].font = font
        dns_sheet_empty_row += 1
        for views in view_dns:
            dns_sheet['F{}'.format(dns_sheet_empty_row)] = add_nri_cname.format(
                MNC=mnc,
                view_name=views,
                NRI=format(int(ne_info.get('NRI')), '02x').upper(),
                RAC=format(int(ne_info.get('LAI')[x].get('RAC')), '02x').upper(),
                LAC=format(int(ne_info.get('LAI')[x].get('LAC')), '04x').upper(),
                SGSN=ne_info.get('SGSN Name')
            )
            dns_sheet_empty_row += 1
        dns_sheet_empty_row += 2
        dns_sheet['F{}'.format(dns_sheet_empty_row)] = 'SAV CFG:;'
        dns_sheet_empty_row += 2
    """
    # Fallback iDNS Sheet
    fallback_dns_sheet['F{}'.format(fallback_dns_sheet_empty_row)] = '/* Projeto DNS F5 */'
    fallback_dns_sheet['F{}'.format(fallback_dns_sheet_empty_row)].font = font
    fallback_dns_sheet_empty_row += 2
    fallback_dns_sheet['F{}'.format(fallback_dns_sheet_empty_row)] = '{}'.format(fallback_project_name)
    fallback_dns_sheet_empty_row += 1
    """
    Script Deprecreated due to SWAP of Huawei DNS
    for x in xrange(lai_count):
        mnc = mnc_select.get(ne_info.get('LAI')[x].get('LAC')[-2:-1])
        fallback_dns_sheet['F{}'.format(fallback_dns_sheet_empty_row)] = '/* LAC: {}  RAC: {} */'.format(
            ne_info.get('LAI')[x].get('LAC'),
            ne_info.get('LAI')[x].get('RAC'))
        fallback_dns_sheet['F{}'.format(fallback_dns_sheet_empty_row)].font = font
        fallback_dns_sheet_empty_row += 1
        if not expansion:
            for views in view_dns:
                fallback_dns_sheet['F{}'.format(fallback_dns_sheet_empty_row)] = rmv_rac_cname.format(
                    MNC=mnc,
                    view_name=views,
                    RAC=format(int(ne_info.get('LAI')[x].get('RAC')), '02x').upper(),
                    LAC=format(int(ne_info.get('LAI')[x].get('LAC')), '04x').upper()
                )
                fallback_dns_sheet_empty_row += 1
        fallback_dns_sheet_empty_row += 2
        fallback_dns_sheet['F{}'.format(fallback_dns_sheet_empty_row)] = '/* NRI: {} */'.format(ne_info.get('NRI'))
        fallback_dns_sheet['F{}'.format(fallback_dns_sheet_empty_row)].font = font
        fallback_dns_sheet_empty_row += 1
        for views in view_dns:
            fallback_dns_sheet['F{}'.format(fallback_dns_sheet_empty_row)] = rmv_nri_cname.format(
                MNC=mnc,
                view_name=views,
                NRI=format(int(ne_info.get('NRI')), '02x').upper(),
                RAC=format(int(ne_info.get('LAI')[x].get('RAC')), '02x').upper(),
                LAC=format(int(ne_info.get('LAI')[x].get('LAC')), '04x').upper()
            )
            fallback_dns_sheet_empty_row += 1
        fallback_dns_sheet_empty_row += 2
        dns_sheet['F{}'.format(fallback_dns_sheet_empty_row)] = 'SAV CFG:;'
        fallback_dns_sheet_empty_row += 2
    """
    return dns_sheet_empty_row, fallback_dns_sheet_empty_row


def create_bsc_workbook(bsc_data, full_path):
    date = datetime.datetime.now()

    destination_path = os.path.join(full_path,
                                    'Projeto Integracao BSC_{}_{}.xlsx'.format(bsc_data.get('BSC Name'),
                                                                               date.strftime('%d%m%Y'))
                                    )
    if not os.path.exists(destination_path):
        base_folder = os.path.abspath(os.getcwd())
        excel_template = os.path.join(base_folder, 'excel_templates\Project_Template.xlsx')
        try:
            wb = pyxl.load_workbook(filename=excel_template)
            wb.save(filename=destination_path)
            return destination_path
        except exceptions.IOError as error:
            if error[0] == 2:
                print('Sorry, but the path to excel_templates does not exist.')
                print('Path expected: "{}"'.format(excel_template))
            else:
                print('Sorry, there was an IOError.')
                print('ERROR: {}'.format(error))
    else:
        return destination_path


def create_bsc_project(workbook_destination_path, user_name, bsc_data, file_input, fallback_empty_row,
                       dns_sheet_empty_row, fallback_dns_sheet_empty_row, expansion, sgsn_bsc_count):
    digit_pattern = "[-+]?\d+[\.]?\d*"  # Regex para encontrar digitos numa string
    wb_from = pyxl.load_workbook(file_input, keep_vba=True)  # Abrir o Gb_Iu_Planning mantendo o codigo VBA
    ws_from = wb_from.get_sheet_by_name('GBoIP')  # Selecionar a Aba "GBoIP"
    # Abrir o workbook da BSC
    workbook = pyxl.load_workbook(workbook_destination_path)
    # Criar aba com nome do SGSN
    ws_to = workbook.create_sheet(title='Projeto {}'.format(bsc_data.get('SGSN Name')), index=2)
    ws_to_fallback = workbook.get_sheet_by_name('Fallback')  # Selecionar aba Fallback
    merged = ws_from.merged_cell_ranges  # Encontrar todas as Células Mescladas
    merged = sorted(merged, key=lambda x: byord(x))  # Ordernar a lista com base nas colunas
    merged = sorted(merged, key=lambda x: bydigit(x))  # Ordernar a lista com base nas linhas

    # Copiar os cabeçalhos do Gb_Iu_Planning para a nova aba do novo Workbook
    for rrow in ws_from.iter_rows('A1:Q3'):
        for cell in rrow:
            ws_to[cell.coordinate].value = cell.value
            ws_to[cell.coordinate].style = cell.style

    # Mesclar as células do cabeçalho
    for merged_cells in merged[:18]:
        ws_to.merge_cells(merged_cells)

    # Copiar as células com informações da BSC alvo para a nova aba do novo Workbook
    row_count = 0
    for rrow in ws_from.iter_rows(
            'A{}:Q{}'.format(bsc_data.get('Coordinate')[0], bsc_data.get('Coordinate')[1])):
        for cell in rrow:
            ranges = re.findall(digit_pattern, cell.coordinate)  # Encontrar as linhas correspondentes das celulas

            coordinate_normalized = 4 + row_count

            # Criar a nova coordenada já normalizada
            cell_coordinate = cell.coordinate.replace(ranges[0], str(coordinate_normalized))
            # Copiar os valores das células
            ws_to[cell_coordinate] = cell.value
            # Copiar os estilos das células
            ws_to[cell_coordinate].style = cell.style
        row_count += 1

    # Encontrar o INDEX da célula BSC Name da BSC alvo
    start_row = bsc_data.get('Coordinate')[0]
    end_row = bsc_data.get('Coordinate')[1]
    bsc_name_index = merged.index('A{}:A{}'.format(start_row, end_row))

    # Mesclar as células que contém informação da BSC alvo

    for merged_cells in merged[bsc_name_index:bsc_name_index + 6]:
        ranges = re.findall(digit_pattern, merged_cells)  # Encontrar as linhas correspondentes das celulas
        # Normalizar a coordenada para iniciar a Mescla a partir da célula A3
        first_range_normalized = 4
        # Normalizar a coordenada para finalizar a Mescla na célula correta
        second_range_normalized = first_range_normalized + (int(ranges[1]) - int(ranges[0]))

        temp_coordinate = merged_cells.replace(ranges[0], str(first_range_normalized))
        final_coordinate = temp_coordinate.replace(ranges[1], str(second_range_normalized))
        # Mesclar as células com as coordenadas normalizadas
        ws_to.merge_cells(final_coordinate)

    fulfill_cover_info(workbook, 'Projeto Integração BSC - {}'.format(bsc_data.get('BSC Name')), user_name)
    fallback_empty_row = fulfill_bsc_info(ws_to, ws_to_fallback, bsc_data, fallback_empty_row)
    if sgsn_bsc_count == 1:
        dns_sheet_empty_row, fallback_dns_sheet_empty_row = fulfill_dns_info(workbook, bsc_data,
                                                                             dns_sheet_empty_row,
                                                                             fallback_dns_sheet_empty_row,
                                                                             expansion)
    # Salvar o novo Workbook
    workbook.save(filename=workbook_destination_path)

    return fallback_empty_row, dns_sheet_empty_row, fallback_dns_sheet_empty_row


def fulfill_bsc_info(bsc_worksheet, fallback, bsc_info, fallback_empty_row):
    # Define font style
    font = pyxlst.Font(name='Arial', size=11, bold=True, underline="single")

    # BSC Sheet

    # Configure NSEID Destination
    empty_row = bsc_worksheet.max_row + 3
    bsc_worksheet['A{}'.format(empty_row)] = '/* To use on BATCH COMMAND */'
    bsc_worksheet['A{}'.format(empty_row)].font = font
    empty_row = bsc_worksheet.max_row + 1
    bsc_worksheet['A{}'.format(empty_row)] = 'USE ME:MEID = 8;'

    empty_row = bsc_worksheet.max_row + 2
    bsc_worksheet['A{}'.format(empty_row)] = '/* 1 - Add an destination NSEID */'
    bsc_worksheet['A{}'.format(empty_row)].font = font
    empty_row = bsc_worksheet.max_row + 1
    nsei_count = len(bsc_info.get('NSEI'))
    for x in xrange(nsei_count):
        bsc_worksheet['A{}'.format(empty_row + x)] = add_nse.format(BSC_Name=bsc_info.get('BSC Name'),
                                                                    nsei=bsc_info.get('NSEI')[x],
                                                                    srn=bsc_info.get('GBLOCALENDPOINT')[x].get('SRN'),
                                                                    sn=bsc_info.get('GBLOCALENDPOINT')[x].get('SN'),
                                                                    tx=bsc_info.get('TX'),
                                                                    NSE_Type=bsc_info.get('NSE Conf Type'),
                                                                    GB_Flex=bsc_info.get('GB Flex'))

    # Add local link NSEID with local IP address
    empty_row = bsc_worksheet.max_row + 2
    bsc_worksheet['A{}'.format(empty_row)] = '/* 2 - Add local link NSEID with local IP address */'
    bsc_worksheet['A{}'.format(empty_row)].font = font
    empty_row = bsc_worksheet.max_row + 1
    for x in xrange(nsei_count):
        bsc_worksheet['A{}'.format(empty_row + x)] = add_gbiplocendpt.format(nsei=bsc_info.get('NSEI')[x],
                                                                             srn=bsc_info.get(
                                                                                 'GBLOCALENDPOINT')[x].get('SRN'),
                                                                             sn=bsc_info.get(
                                                                                 'GBLOCALENDPOINT')[x].get('SN'),
                                                                             LIPV4=bsc_info.get(
                                                                                 'GBLOCALENDPOINT')[x].get('LOCALIP1'),
                                                                             LOCALPORT=bsc_info.get(
                                                                                 'GBLOCALENDPOINT')[x].get('LOCALPORT'),
                                                                             BSC_Name=bsc_info.get('BSC Name'))

        bsc_worksheet['A{}'.format(empty_row + x + 1)] = add_gbiplocendpt.format(nsei=bsc_info.get('NSEI')[x],
                                                                                 srn=bsc_info.get(
                                                                                     'GBLOCALENDPOINT')[x].get('SRN'),
                                                                                 sn=bsc_info.get(
                                                                                     'GBLOCALENDPOINT')[x].get('SN'),
                                                                                 LIPV4=bsc_info.get(
                                                                                     'GBLOCALENDPOINT')[x].get(
                                                                                     'LOCALIP2'),
                                                                                 LOCALPORT=bsc_info.get(
                                                                                     'GBLOCALENDPOINT')[x].get(
                                                                                     'LOCALPORT'),
                                                                                 BSC_Name=bsc_info.get('BSC Name'))

    # Fallback Sheet
    fallback['E{}'.format(fallback_empty_row)] = bsc_info.get('SGSN Name')
    fallback['E{}'.format(fallback_empty_row)].font = font
    fallback_empty_row += 1
    for x in xrange(nsei_count):
        fallback['E{}'.format(fallback_empty_row)] = rmv_gbiplocendpt.format(nsei=bsc_info.get('NSEI')[x],
                                                                             LIPV4=bsc_info.get(
                                                                                 'GBLOCALENDPOINT')[x].get('LOCALIP1'),
                                                                             LOCALPORT=bsc_info.get(
                                                                                 'GBLOCALENDPOINT')[x].get('LOCALPORT')
                                                                             )

        fallback_empty_row += 1

        fallback['E{}'.format(fallback_empty_row)] = rmv_gbiplocendpt.format(nsei=bsc_info.get('NSEI')[x],
                                                                             LIPV4=bsc_info.get(
                                                                                 'GBLOCALENDPOINT')[x].get('LOCALIP2'),
                                                                             LOCALPORT=bsc_info.get(
                                                                                 'GBLOCALENDPOINT')[x].get('LOCALPORT')
                                                                             )
        fallback_empty_row += 1

    fallback_empty_row += 1

    fallback['E{}'.format(fallback_empty_row)] = rmv_nse.format(BSC_Name=bsc_info.get('BSC Name'))

    fallback_empty_row += 2

    return fallback_empty_row


def srvcc_dns_entries_f5(ne_info, expansion):
    lai_count = len(ne_info.get('LAI'))
    create_naptrrecord_f5 = list()
    lac_list = list()
    if not expansion:
        for x in xrange(lai_count):
            lac_list.append(ne_info.get('LAI')[x].get('LAC'))
        lac_list = set(lac_list)

        for lac in lac_list:
            for msc in srvcc_msc.get(lac[-2:]):
                mnc = mnc_select.get(lac[-2:-1])
                create_naptrrecord_f5.append({
                    'action': 'add',
                    'domain_name': 'lac{LAC}.lac.epc.mnc{mnc}.mcc724.3gppnetwork.org.'.format(
                        LAC=format(int(lac), '04x').upper(),
                        mnc=mnc
                    ),
                    'order': 10,
                    'preference': 10,
                    'flags': 'A',
                    'service': 'x-3gpp-msc:x-sv',
                    'regexp': '""',
                    'replacement': 'topoff.vip-sv.{node_name}.node.epc.mnc{mnc}.mcc724.3gppnetwork.org.'.format(
                        node_name=msc,
                        mnc=mnc
                    ),
                    'ttl': 300}
                )

    return create_naptrrecord_f5


def export_f5_entries(ne_info, full_path, entries_list):
    date = datetime.datetime.now()
    date_string = '{}'.format(date.strftime('%d%m%Y'))
    ne_name = ne_info.get('RNC Name')
    destination_path = os.path.join(full_path, 'Projeto Integracao RNC_{}_{}.txt'.format(ne_name, date_string)
                                    )
    fallback_destination_path = os.path.join(full_path, 'FALLBACK_Projeto Integracao RNC_{}_{}.txt'.format(ne_name,
                                                                                                           date_string)
                                             )
    if not ne_name:
        ne_name = ne_info.get('BSC Name')
        destination_path = os.path.join(full_path, 'Projeto Integracao BSC_{}_{}.txt'.format(ne_name, date_string)
                                        )
        fallback_destination_path = os.path.join(full_path,
                                                 'FALLBACK_Projeto Integracao BSC_{}_{}.txt'.format(ne_name,
                                                                                                    date_string)
                                                 )

    with open(destination_path, 'wb') as export:
        for entries in entries_list:
            export.write(json.dumps(entries) + '\n')

    with open(fallback_destination_path, 'wb') as export:
        for entries in entries_list:
            entries['action'] = 'remove'
            export.write(json.dumps(entries) + '\n')


if __name__ == '__main__':
    main()
