# coding=utf-8
import warnings
import openpyxl as pyxl
import openpyxl.styles as pyxlst
import shutil
import os
import datetime
from support_variables import bayface, nsei_start_number
from support_functions import find_between, roundrobin, export_excel_to_csv, export_csv_to_dict_list

warnings.filterwarnings("ignore")
slot_count_0 = 0
slot_count_1 = 0

# TODO implement OrderedDict to try to solve NSEI and Local IPs relations in excel (populate_gbplanning_with_bsc_info)
alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
            'U', 'V', 'W', 'X', 'Y', 'Z']
rnc_column_id = {
    'RNC Name': 'A',
    'Vendor': 'B',
    'RNCX': 'C',
    'RNCID': 'D',
    'DPC': 'E',
    'DPX': 'F',
    'RNCMCC': 'G',
    'RNCMNC': 'H',
    'NI': 'I',
    'Iu-flex': 'J',
    'LAC': 'K',
    'RAC': 'L',
    'RanSharing': 'M',
    'DT': 'N',
    'Support R7 Qos': 'O',
    'RAB QOS': 'P',
    'RNC Version': 'Q',
    'SGSN Name': 'R',
    'NRI': 'S',
    'OPC': 'T',
    'OPX': 'U',
    'NI Logical Link': 'V',
    'ESU Subrack': 'W',
    'ESU Slot': 'X',
    'LNK': 'Y',
    'PRI': 'Z',
    'Entity Type': 'AA',
    'RNC Entity Type': 'AB',
    'LEX': 'AC',
    'DEX': 'AD',
    'LSX': 'AE',
    'DET': 'AF',
    'SGSN IP1': 'AG',
    'SGSN IP2': 'AH',
    'SGSN Port': 'AI',
    'RNC IP1': 'AJ',
    'RNC IP2': 'AK',
    'RNC Port': 'AL',
    'Share Type': 'AM',
    'C/S': 'AN',
    'SSNX (RANAP)': 'AO',
    'SSNX (SCMG)': 'AP',
    'SLSM': 'AQ'
}
bsc_column_id = {
    'BSC Name': 'A',
    'TX': 'B',
    'NSEI': 'C',
    'BSSID': 'D',
    'NSE Conf Type': 'E',
    'SGSN Name': 'F',
    'NRI': 'G',
    'Subrack No': 'H',
    'Slot No': 'I',
    'Local IP Address 1': 'J',
    'Local IP Address 2': 'K',
    'Local Port': 'L',
    'GB-FLEX': 'M'}


def main():
    print(u'Welcome to the Gb Iu Planning creation Tool!')
    user_name = raw_input(u'Please, type in your full name: ')
    while True:
        print(u'{}, choose one of the options below to start: '.format(user_name.split(' ')[0]))
        print(u'1: Update Gb Iu Planning from legacy USN')
        print(u'2: Create new Gb Iu Planning from legacy USN')
        print(u'3: Exit')
        choice = raw_input('> ').strip()
        if choice == '1':
            legacy_usn = raw_input(u'Type in the path to the legacy USN export file: ').strip()
            base_folder = os.path.abspath(os.getcwd())
            destination_folder = 'Gb_Iu_Planning'
            full_path = os.path.join(base_folder, destination_folder)
            if not os.path.exists(full_path):
                os.makedirs(full_path)
            legacy_usn_data = extract_usn_info(legacy_usn)
            all_rnc_data = extract_rnc_info(legacy_usn)
            bsc_data = extract_bsc_info(legacy_usn)
            create_gb_iu_planning(full_path, user_name, legacy_usn_data, all_rnc_data, bsc_data)
            print('**************************')
            print('* Gb Iu Planning Created *')
            print('**************************')
        elif choice == '2':
            legacy_usn = raw_input(u'Type in the path to the legacy USN export file: ').strip()
            new_usn = raw_input(u'Type in the path to the new USN export file: ').strip()
            base_folder = os.path.abspath(os.getcwd())
            destination_folder = 'Gb_Iu_Planning'
            full_path = os.path.join(base_folder, destination_folder)
            if not os.path.exists(full_path):
                os.makedirs(full_path)
            legacy_usn_data = extract_usn_info(legacy_usn)
            new_usn_data = extract_usn_info(new_usn)
            all_rnc_data = extract_rnc_info(legacy_usn)
            bsc_data = extract_bsc_info(legacy_usn)
            create_gb_iu_planning(full_path, user_name, new_usn_data, all_rnc_data, bsc_data, new_usn=True,
                                  legacy_usn_data=legacy_usn_data)
            print('**************************')
            print('* Gb Iu Planning Created *')
            print('**************************')
        elif choice == '3':
            print(u'Thank you for using the Gb Iu Planning creation Tool!')
            print(u'If you have any suggestions please send an e-mail to: decastromonteiro@gmail.com')
            raw_input()
            break

        else:
            print(u"Sorry, we don't understand ['{}'], please try again.".format(choice))
            continue


def extract_usn_info(file_input):
    usn_data = dict()
    local_ipv4_iuc = list()
    local_ipv4_gb = list()

    with open(file_input) as fin:
        for line in fin:
            line = line.strip()
            if line.startswith('ADD ME:'):
                if find_between(line, 'MEID=', ',') == '8':
                    usn_data['USN'] = find_between(line, 'MENAME="', '",')
            elif line.startswith('SET SYS:'):
                usn_data['NRI'] = find_between(line, 'CNID=', ',')
            elif line.startswith('ADD BRDIP:'):
                if 'IU_C' in find_between(line, 'DESC="', '";'):
                    local_ipv4_iuc.append(
                        {'SRN': find_between(line, 'SRN=', ','),
                         'IP': find_between(line, 'IPV4="', '",')}
                    )

                elif 'Gb' in find_between(line, 'DESC="', '";'):
                    local_ipv4_gb.append(
                        {'SRN': find_between(line, 'SRN=', ','),
                         'IP': find_between(line, 'IPV4="', '",')}
                    )
                usn_data['Iu_C'] = local_ipv4_iuc
                usn_data['Gb'] = local_ipv4_gb
            elif line.startswith('ADD M3LE:'):
                usn_data['OPC'] = find_between(line, 'OPC="0x', '",').upper()

    return usn_data


def extract_rnc_info(file_input):
    all_rnc_data = list()

    fin = open(file_input)
    line_collection = fin.readlines()

    for line in line_collection:
        rnc_info = dict()
        m3_lnk = list()
        lai = list()
        line = line.strip()
        if line.startswith('ADD M3DE:'):
            if find_between(line, 'DET=', ',') == 'IPSP':
                rnc_info['DEX'] = find_between(line, 'DEX=', ',')
                rnc_info['LEX'] = find_between(line, 'LEX=', ',')
                rnc_info['DPC'] = find_between(line, 'DPC="0x', '",').upper()
                rnc_info['DET'] = 'IPSP'
                rnc_info['RNCN'] = find_between(line, 'DEN="', '";')
                for nextline in line_collection:
                    nextline.strip()
                    if nextline.startswith('ADD M3LKS:'):
                        if rnc_info.get('RNCN') == find_between(nextline, 'LSN="', '";'):
                            rnc_info['LSX'] = find_between(nextline, 'LSX=', ',')
                            rnc_info['SLSM'] = find_between(nextline, 'SLSM=', ',')
                            rnc_info['TM'] = find_between(nextline, 'TM=', ',')
                    elif nextline.startswith('ADD M3RT:'):
                        if rnc_info.get('RNCN') == find_between(nextline, 'RTN="', '";'):
                            rnc_info['RTX'] = find_between(nextline, 'RTX=', ',')
                    elif nextline.startswith('ADD M3LNK:'):
                        if rnc_info.get('RNCN') in find_between(nextline, 'LKN="', '";'):
                            m3_lnk.append(
                                {'SRN': find_between(nextline, 'SRN=', ','),
                                 'SN': find_between(nextline, 'SN=', ','),
                                 'LNK': find_between(nextline, 'LNK=', ','),
                                 'LOCIPV41': find_between(nextline, 'LOCIPV41="', '",'),
                                 'LOCIPV42': find_between(nextline, 'LOCIPV42="', '",'),
                                 'LOCPORT': find_between(nextline, 'LOCPORT=', ','),
                                 'PEERIPV41': find_between(nextline, 'PEERIPV41="', '",'),
                                 'PEERIPV42': find_between(nextline, 'PEERIPV42="', '",'),
                                 'PEERPORT': find_between(nextline, 'PEERPORT=', ','),
                                 'SCTPINDX': find_between(nextline, 'SCTPINDX=', ',')
                                 }
                            )
                            rnc_info['M3LNK'] = m3_lnk
                    elif nextline.startswith('ADD SCCPDPC:'):
                        if rnc_info.get('RNCN') == find_between(nextline, 'DPN="', '";'):
                            rnc_info['DPX'] = find_between(nextline, 'DPX=', ',')
                    elif nextline.startswith('ADD SCCPSSN:'):
                        if rnc_info.get('RNCN') == find_between(nextline, 'SSNNAME="', '",'):
                            if find_between(nextline, 'SSN=', ',') == 'RANAP':
                                rnc_info['SSNX_RANAP'] = find_between(nextline, 'SSNX=', ',')
                            elif find_between(nextline, 'SSN=', ',') == 'SCMG':
                                rnc_info['SSNX_SCMG'] = find_between(nextline, 'SSNX=', ',')
                    elif nextline.startswith('ADD RNC:'):
                        if rnc_info.get('RNCN') == find_between(nextline, 'RNCN="', '",'):
                            rnc_info['RNCX'] = find_between(nextline, 'RNCX=', ',')
                            rnc_info['RNCMCC'] = find_between(nextline, 'RNCMCC="', '",')
                            rnc_info['RNCMNC'] = find_between(nextline, 'RNCMNC="', '",')
                            rnc_info['RNCID'] = find_between(nextline, 'RNCID=', ',')
                            rnc_info['NI'] = find_between(nextline, 'NI=', ',')
                            rnc_info['SPC'] = find_between(nextline, 'SPC="0x', '",').upper()
                            rnc_info['CNID'] = find_between(nextline, 'CNID=', ',')
                            rnc_info['IMS'] = find_between(nextline, 'IMS=', ',')
                            rnc_info['IU-FLEX'] = find_between(nextline, 'IU-FLEX=', ',')
                            rnc_info['RABQOS'] = find_between(nextline, 'RABQOS=', ',')
                    elif nextline.startswith('ADD IUPAGING:'):
                        if rnc_info.get('RNCX') == find_between(nextline, 'RNCINDEX=', ';'):
                            plmn = rnc_info.get('RNCMCC') + rnc_info.get('RNCMNC')
                            lai.append(
                                {'LAC': str(int(find_between(nextline, 'LAI="{}'.format(plmn), '",'), 16)),
                                 'RAC': str(int(find_between(nextline, 'RAC="0x', '",'), 16))}
                            )
                            rnc_info['LAI'] = lai
        if rnc_info:
            all_rnc_data.append(rnc_info)

    all_rnc_data = sorted(all_rnc_data, key=lambda k: k['RNCX'])
    return all_rnc_data


def extract_bsc_info(file_input):

    def populate_local_endpoint(bsc_info, gblocalendpoint_list):
        nsei_count = len(bsc_info.get('NSE'))
        local_endpoint = list()
        for x in xrange(nsei_count):
            for nsei_dict in gblocalendpoint_list:
                endpoint = nsei_dict.get(bsc_info.get('NSE')[x].get('NSEI'))
                if endpoint:
                    local_endpoint.append(endpoint)
        return local_endpoint

    def extract_gplocalendpoint_info():
        gplocalendpoint = list()
        with open(file_input) as fin:
            for lines in fin:
                if lines.startswith('ADD GBIPLOCENDPT'):
                    gplocalendpoint.append({
                        find_between(lines, 'NSEI=', ','): {
                            'SRN': find_between(lines, 'SRN=', ','),
                            'SN': find_between(lines, 'SN=', ','),
                            'LIPV4': find_between(lines, 'LIPV4="', '",'),
                            'LUP': find_between(lines, 'LUP=', ',')
                        }
                    })
        return gplocalendpoint

    gblocalendpoint_list = extract_gplocalendpoint_info()

    all_bsc_data = list()

    fin = open(file_input)
    line_collection = fin.readlines()
    line_collection.sort(reverse=True)
    count_line = 0
    for line in line_collection:

        bsc_info = dict()
        nse = list()
        line = line.strip()
        count = len(all_bsc_data) - 1
        if line.startswith('ADD NSE'):
            if count == -1:
                bsc_info['BSC Name'] = extract_bsc_name_due_to_dumbness(find_between(line, 'OTHERNODE="', '",'))
                nse.append(
                    {'NSEI': find_between(line, 'NSEI=', ','),
                     'SRN': find_between(line, 'SRN=', ','),
                     'SN': find_between(line, 'SN=', ','),
                     'BSSID': find_between(line, 'BSSID=', ',')
                     }
                )
                bsc_info['NSE'] = nse
                bsc_info['PFC'] = find_between(line, 'PFC=', ',')
                bsc_info['BT'] = find_between(line, 'BT=', ',')
                bsc_info['CT'] = find_between(line, 'CT=', ',')
                bsc_info['GB-FLEX'] = find_between(line, 'GB-FLEX=', ',')
                for line0 in line_collection[count_line + 1:]:
                    if line0.startswith('ADD NSE'):
                        if bsc_info['BSC Name'] == extract_bsc_name_due_to_dumbness(
                                find_between(line0, 'OTHERNODE="', '",')):
                            nse.append(
                                {'NSEI': find_between(line0, 'NSEI=', ','),
                                 'SRN': find_between(line0, 'SRN=', ','),
                                 'SN': find_between(line0, 'SN=', ','),
                                 'BSSID': find_between(line0, 'BSSID=', ',')
                                 }
                            )
                        bsc_info['NSE'] = nse

            else:
                if not all_bsc_data[count].get('BSC Name') == extract_bsc_name_due_to_dumbness(find_between(
                        line, 'OTHERNODE="', '",')):

                    bsc_info['BSC Name'] = extract_bsc_name_due_to_dumbness(find_between(line, 'OTHERNODE="', '",'))
                    nse.append(
                        {'NSEI': find_between(line, 'NSEI=', ','),
                         'SRN': find_between(line, 'SRN=', ','),
                         'SN': find_between(line, 'SN=', ','),
                         'BSSID': find_between(line, 'BSSID=', ',')
                         }
                    )
                    bsc_info['NSE'] = nse
                    bsc_info['PFC'] = find_between(line, 'PFC=', ',')
                    bsc_info['BT'] = find_between(line, 'BT=', ',')
                    bsc_info['CT'] = find_between(line, 'CT=', ',')
                    bsc_info['GB-FLEX'] = find_between(line, 'GB-FLEX=', ',')
                    nsei_counter = 0
                    for line0 in line_collection[count_line + 1:]:
                        if line0.startswith('ADD NSE'):
                            if bsc_info['BSC Name'] == extract_bsc_name_due_to_dumbness(
                                    find_between(line0, 'OTHERNODE="', '",')):
                                nse.append(
                                    {'NSEI': find_between(line0, 'NSEI=', ','),
                                     'SRN': find_between(line0, 'SRN=', ','),
                                     'SN': find_between(line0, 'SN=', ','),
                                     'BSSID': find_between(line0, 'BSSID=', ',')
                                     }
                                )
                                nsei_counter += 1
                            bsc_info['NSE'] = nse

        if bsc_info:
            bsc_info['GBLOCALENDPOINT'] = populate_local_endpoint(bsc_info, gblocalendpoint_list)
            all_bsc_data.append(bsc_info)
        count_line += 1
    all_bsc_data = sorted(all_bsc_data, key=lambda k: k['BSC Name'])
    return all_bsc_data


def extract_bsc_name_due_to_dumbness(bscname):
    if bscname.startswith('TO'):
        bscname = bscname[3:]
        return bscname
    else:
        return bscname


def create_gb_iu_planning(base_folder, user_name, usn_data, rnc_data, bsc_data, new_usn=False, legacy_usn_data=None):
    port_number = 2905  # Initial port_number to populate new Iu_C sheet of USN Gb_Iu_Planning
    port_number_gb = int('5{}000'.format(nsei_start_number.get(usn_data.get('USN'))[0]))
    esu_subrack = ['0', '1', '0', '1']  # Initial esu_subrack list to help choosing Slot Numbers
    date = datetime.datetime.now()  # Today date information
    template_path = os.path.abspath('./excel_templates/Gb_Iu_Planning_Template.xlsx')  # Path to Excel Template
    gb_iu_planning_path = os.path.join(base_folder, 'Gb_Iu_Planning_{}_{}.xlsx'.format(
        usn_data.get('USN'), date.strftime('%d%m%Y'))  # Path to new Gb_Iu_Planning Workbook
                                       )
    shutil.copyfile(template_path, gb_iu_planning_path)  # Create new Gb_Iu_Planning using the Excel Template
    active_workbook = pyxl.load_workbook(gb_iu_planning_path)  # Load new Gb_Iu_Planning to work with it

    # Aba Capa
    cover = active_workbook.get_sheet_by_name('Capa')
    cover['F9'] = date.strftime('%d/%m/%Y')
    cover['J6'] = 'Gb Iu Planning do Elemento {}'.format(usn_data.get('USN'))
    cover['H13'] = user_name.strip()
    cover['D21'] = user_name.strip()
    cover['H21'] = date.strftime('%d/%m/%Y')
    cover['K21'] = '1'
    cover['M21'] = 'Gb Iu Planning atualizado.'

    # Set global alignment variable to format cells
    alignment = pyxlst.Alignment(horizontal='center',
                                 vertical='center',
                                 text_rotation=0,
                                 wrap_text=False,
                                 shrink_to_fit=False,
                                 indent=0)
    # Set global border variable to format cells
    border = pyxlst.Border(left=pyxlst.Side(border_style='thin', color='000000'),
                           right=pyxlst.Side(border_style='thin', color='000000'),
                           top=pyxlst.Side(border_style='thin', color='000000'),
                           bottom=pyxlst.Side(border_style='thin', color='000000'),
                           )
    # Aba Iu_C
    iu_c = active_workbook.get_sheet_by_name('Iu_C')
    for counter in xrange(len(rnc_data)):
        # Calculate the maximum number of rows used for one RNC
        rows_num = max(len(rnc_data[counter].get('LAI')), len(rnc_data[counter].get('M3LNK')))
        if counter == 0:
            empty_row = iu_c.max_row + 1  # Calculate first empty row on the Excel Sheet
            end_row = empty_row + rows_num - 1  # Calculate last row after fulfilling information
            if new_usn:
                populate_iuc_sheet_with_usn_info(iu_c, empty_row, end_row, rnc_data, counter,
                                                 usn_data, esu_subrack, port_number)
            else:
                populate_iuc_sheet_with_rnc_info(iu_c, empty_row, end_row, rnc_data, counter, usn_data)
            # Format first cell with RNC Name
            first_cell = iu_c['{}{}'.format(rnc_column_id.get('RNC Name'), empty_row)]
            fill = pyxlst.PatternFill(fill_type="solid", fgColor='32CD32')
            first_cell.fill = fill

        else:
            empty_row = end_row + 1  # Calculate first empty row on the Excel Sheet after first information fulfilled
            end_row = empty_row + rows_num - 1  # Calculate last row after fulfilling information
            if new_usn:
                populate_iuc_sheet_with_usn_info(iu_c, empty_row, end_row, rnc_data, counter,
                                                 usn_data, esu_subrack, port_number)
            else:
                populate_iuc_sheet_with_rnc_info(iu_c, empty_row, end_row, rnc_data, counter, usn_data)
            # Format every cell with RNC Name
            first_cell = iu_c['{}{}'.format(rnc_column_id.get('RNC Name'), empty_row)]
            fill = pyxlst.PatternFill(fill_type="solid", fgColor='32CD32')
            first_cell.fill = fill
        port_number += 20

    # Aba GBoIP
    gbo_ip = active_workbook.get_sheet_by_name('GBoIP')
    for counter in xrange(len(bsc_data)):
        try:
            rows_num = len(bsc_data[counter].get('NSE'))
        except TypeError:
            rows_num = len(bsc_data[counter].get('NSE'))

        if counter == 0:

            if new_usn:
                if bsc_data[counter].get('GB-FLEX') == 'YES':
                    empty_row_gb = gbo_ip.max_row + 1
                    end_row_gb = empty_row_gb + rows_num - 1

                    port_number_gb = populate_gboip_sheet_with_usn_info(gbo_ip, empty_row_gb, end_row_gb, bsc_data,
                                                                        counter, usn_data, esu_subrack, port_number_gb,
                                                                        legacy_usn_data)

            else:
                empty_row_gb = gbo_ip.max_row + 1
                end_row_gb = empty_row_gb + rows_num - 1
                populate_gboip_sheet_with_bsc_info(gbo_ip, empty_row_gb, end_row_gb, bsc_data, counter, usn_data)
            # Format first cell with BSC Name
            first_cell = gbo_ip['{}{}'.format(bsc_column_id.get('BSC Name'), empty_row_gb)]
            fill = pyxlst.PatternFill(fill_type="solid", fgColor='32CD32')
            first_cell.fill = fill
        else:

            if not counter % 2:
                esu_subrack = esu_subrack[1:]  # Caso o contador seja impar, começar pelo subrack 1

            if new_usn:
                if bsc_data[counter].get('GB-FLEX') == 'YES':
                    empty_row_gb = end_row_gb + 1
                    end_row_gb = empty_row_gb + rows_num - 1
                    port_number_gb = populate_gboip_sheet_with_usn_info(gbo_ip, empty_row_gb, end_row_gb, bsc_data,
                                                                        counter, usn_data, esu_subrack, port_number_gb,
                                                                        legacy_usn_data)
            else:
                empty_row_gb = end_row_gb + 1
                end_row_gb = empty_row_gb + rows_num - 1
                populate_gboip_sheet_with_bsc_info(gbo_ip, empty_row_gb, end_row_gb, bsc_data, counter, usn_data)
            # Format every cell with BSC Name
            first_cell = gbo_ip['{}{}'.format(bsc_column_id.get('BSC Name'), empty_row_gb)]
            fill = pyxlst.PatternFill(fill_type="solid", fgColor='32CD32')
            first_cell.fill = fill

    # Format every used cell on Iu_C sheet
    for letters in alphabet:
        for i in xrange(1, end_row + 1):
            active_cell = iu_c[letters + str(i)]
            active_cell.alignment = alignment
            active_cell.border = border
        for other_letters in alphabet[:17]:
            for j in xrange(1, end_row + 1):
                active_cell_sec = iu_c['A' + other_letters + str(j)]
                active_cell_sec.alignment = alignment
                active_cell_sec.border = border

    # Format every used cell on GBoIP sheet
    for letras in alphabet[:17]:
        for i in xrange(1, end_row_gb + 1):
            active_cell = gbo_ip[letras + str(i)]
            active_cell.alignment = alignment
            active_cell.border = border

    active_workbook.save(gb_iu_planning_path)  # Save newly created Workbook
    return gb_iu_planning_path  # Return Workbook Path


def populate_iuc_sheet_with_rnc_info(iu_c, empty_row, end_row, rnc_data, counter, usn_data):
    iu_c.merge_cells(start_row=empty_row, start_column=1, end_row=end_row, end_column=1)
    iu_c.merge_cells(start_row=empty_row, start_column=2, end_row=end_row, end_column=2)
    iu_c.merge_cells(start_row=empty_row, start_column=3, end_row=end_row, end_column=3)
    iu_c.merge_cells(start_row=empty_row, start_column=4, end_row=end_row, end_column=4)
    iu_c.merge_cells(start_row=empty_row, start_column=5, end_row=end_row, end_column=5)
    iu_c.merge_cells(start_row=empty_row, start_column=6, end_row=end_row, end_column=6)
    iu_c.merge_cells(start_row=empty_row, start_column=7, end_row=end_row, end_column=7)
    iu_c.merge_cells(start_row=empty_row, start_column=8, end_row=end_row, end_column=8)
    iu_c.merge_cells(start_row=empty_row, start_column=9, end_row=end_row, end_column=9)
    iu_c.merge_cells(start_row=empty_row, start_column=10, end_row=end_row, end_column=10)
    iu_c.merge_cells(start_row=empty_row, start_column=13, end_row=end_row, end_column=13)
    iu_c.merge_cells(start_row=empty_row, start_column=14, end_row=end_row, end_column=14)
    iu_c.merge_cells(start_row=empty_row, start_column=15, end_row=end_row, end_column=15)
    iu_c.merge_cells(start_row=empty_row, start_column=16, end_row=end_row, end_column=16)
    iu_c.merge_cells(start_row=empty_row, start_column=17, end_row=end_row, end_column=17)
    iu_c.merge_cells(start_row=empty_row, start_column=18, end_row=end_row, end_column=18)
    iu_c.merge_cells(start_row=empty_row, start_column=19, end_row=end_row, end_column=19)
    iu_c.merge_cells(start_row=empty_row, start_column=20, end_row=end_row, end_column=20)
    iu_c.merge_cells(start_row=empty_row, start_column=21, end_row=end_row, end_column=21)
    iu_c.merge_cells(start_row=empty_row, start_column=22, end_row=end_row, end_column=22)
    iu_c.merge_cells(start_row=empty_row, start_column=29, end_row=end_row, end_column=29)
    iu_c.merge_cells(start_row=empty_row, start_column=30, end_row=end_row, end_column=30)
    iu_c.merge_cells(start_row=empty_row, start_column=31, end_row=end_row, end_column=31)
    iu_c.merge_cells(start_row=empty_row, start_column=32, end_row=end_row, end_column=32)
    iu_c.merge_cells(start_row=empty_row, start_column=41, end_row=end_row, end_column=41)
    iu_c.merge_cells(start_row=empty_row, start_column=42, end_row=end_row, end_column=42)
    iu_c.merge_cells(start_row=empty_row, start_column=43, end_row=end_row, end_column=43)
    iu_c['{}{}'.format(rnc_column_id.get('RNC Name'), empty_row)] = rnc_data[counter].get('RNCN')
    iu_c['{}{}'.format(rnc_column_id.get('Vendor'), empty_row)] = 'NA'
    iu_c['{}{}'.format(rnc_column_id.get('RNCX'), empty_row)] = rnc_data[counter].get('RNCX')
    iu_c['{}{}'.format(rnc_column_id.get('RNCID'), empty_row)] = rnc_data[counter].get('RNCID')
    iu_c['{}{}'.format(rnc_column_id.get('DPC'), empty_row)] = rnc_data[counter].get('DPC')
    iu_c['{}{}'.format(rnc_column_id.get('DPX'), empty_row)] = rnc_data[counter].get('DPX')
    iu_c['{}{}'.format(rnc_column_id.get('RNCMCC'), empty_row)] = rnc_data[counter].get('RNCMCC')
    iu_c['{}{}'.format(rnc_column_id.get('RNCMNC'), empty_row)] = rnc_data[counter].get('RNCMNC')
    iu_c['{}{}'.format(rnc_column_id.get('NI'), empty_row)] = 'NATB'
    iu_c['{}{}'.format(rnc_column_id.get('Iu-flex'), empty_row)] = rnc_data[counter].get('IU-FLEX')
    iu_c['{}{}'.format(rnc_column_id.get('RanSharing'), empty_row)] = 'YES'
    iu_c['{}{}'.format(rnc_column_id.get('Iu-flex'), empty_row)] = rnc_data[counter].get('IU-FLEX')
    iu_c['{}{}'.format(rnc_column_id.get('DT'), empty_row)] = 'YES'
    iu_c['{}{}'.format(rnc_column_id.get('Support R7 Qos'), empty_row)] = 'YES'
    iu_c['{}{}'.format(rnc_column_id.get('RAB QOS'), empty_row)] = rnc_data[counter].get('RABQOS')
    iu_c['{}{}'.format(rnc_column_id.get('RNC Version'), empty_row)] = 'R8'
    iu_c['{}{}'.format(rnc_column_id.get('SGSN Name'), empty_row)] = usn_data.get('USN')
    iu_c['{}{}'.format(rnc_column_id.get('NRI'), empty_row)] = usn_data.get('NRI')
    iu_c['{}{}'.format(rnc_column_id.get('OPC'), empty_row)] = usn_data.get('OPC')
    iu_c['{}{}'.format(rnc_column_id.get('OPX'), empty_row)] = '2'
    iu_c['{}{}'.format(rnc_column_id.get('NI Logical Link'), empty_row)] = 'NATB'
    iu_c['{}{}'.format(rnc_column_id.get('LEX'), empty_row)] = rnc_data[counter].get('LEX')
    iu_c['{}{}'.format(rnc_column_id.get('DEX'), empty_row)] = rnc_data[counter].get('DEX')
    iu_c['{}{}'.format(rnc_column_id.get('LSX'), empty_row)] = rnc_data[counter].get('LSX')
    iu_c['{}{}'.format(rnc_column_id.get('DET'), empty_row)] = rnc_data[counter].get('DET')
    iu_c['{}{}'.format(rnc_column_id.get('SSNX (RANAP)'), empty_row)] = rnc_data[counter].get('SSNX_RANAP')
    iu_c['{}{}'.format(rnc_column_id.get('SSNX (SCMG)'), empty_row)] = rnc_data[counter].get('SSNX_SCMG')
    iu_c['{}{}'.format(rnc_column_id.get('SLSM'), empty_row)] = len(rnc_data[counter].get('M3LNK'))
    max_lai = len(rnc_data[counter].get('LAI'))
    lai_count = 0
    for row_count in xrange(empty_row, empty_row + max_lai):
        iu_c['{}{}'.format(rnc_column_id.get('LAC'), row_count)] = rnc_data[counter].get('LAI')[lai_count].get(
            'LAC')
        iu_c['{}{}'.format(rnc_column_id.get('RAC'), row_count)] = rnc_data[counter].get('LAI')[lai_count].get(
            'RAC')
        lai_count += 1

    max_m3lnk = len(rnc_data[counter].get('M3LNK'))
    m3lnk_count = 0
    for row_count in xrange(empty_row, empty_row + max_m3lnk):
        iu_c['{}{}'.format(rnc_column_id.get('ESU Subrack'), row_count)] = rnc_data[counter].get(
            'M3LNK')[m3lnk_count].get('SRN')
        iu_c['{}{}'.format(rnc_column_id.get('ESU Slot'), row_count)] = rnc_data[counter].get(
            'M3LNK')[m3lnk_count].get('SN')
        iu_c['{}{}'.format(rnc_column_id.get('LNK'), row_count)] = rnc_data[counter].get(
            'M3LNK')[m3lnk_count].get('LNK')
        iu_c['{}{}'.format(rnc_column_id.get('PRI'), row_count)] = '0'
        iu_c['{}{}'.format(rnc_column_id.get('Entity Type'), row_count)] = rnc_data[counter].get('DET')
        iu_c['{}{}'.format(rnc_column_id.get('RNC Entity Type'), row_count)] = rnc_data[counter].get('DET')
        iu_c['{}{}'.format(rnc_column_id.get('SGSN IP1'), row_count)] = rnc_data[counter].get(
            'M3LNK')[m3lnk_count].get('LOCIPV41')
        iu_c['{}{}'.format(rnc_column_id.get('SGSN IP2'), row_count)] = rnc_data[counter].get(
            'M3LNK')[m3lnk_count].get('LOCIPV42')
        iu_c['{}{}'.format(rnc_column_id.get('SGSN Port'), row_count)] = rnc_data[counter].get(
            'M3LNK')[m3lnk_count].get('LOCPORT')
        iu_c['{}{}'.format(rnc_column_id.get('RNC IP1'), row_count)] = rnc_data[counter].get(
            'M3LNK')[m3lnk_count].get('PEERIPV41')
        iu_c['{}{}'.format(rnc_column_id.get('RNC IP2'), row_count)] = rnc_data[counter].get(
            'M3LNK')[m3lnk_count].get('PEERIPV42')
        iu_c['{}{}'.format(rnc_column_id.get('RNC Port'), row_count)] = rnc_data[counter].get(
            'M3LNK')[m3lnk_count].get('PEERPORT')
        iu_c['{}{}'.format(rnc_column_id.get('Share Type'), row_count)] = 'LOADSHARE'
        iu_c['{}{}'.format(rnc_column_id.get('C/S'), row_count)] = 'S'
        m3lnk_count += 1


def populate_iuc_sheet_with_usn_info(iu_c, empty_row, end_row, rnc_data, counter, usn_data, esu_subrack, port_number):
    global slot_count_0
    global slot_count_1

    iu_c.merge_cells(start_row=empty_row, start_column=1, end_row=end_row, end_column=1)
    iu_c.merge_cells(start_row=empty_row, start_column=2, end_row=end_row, end_column=2)
    iu_c.merge_cells(start_row=empty_row, start_column=3, end_row=end_row, end_column=3)
    iu_c.merge_cells(start_row=empty_row, start_column=4, end_row=end_row, end_column=4)
    iu_c.merge_cells(start_row=empty_row, start_column=5, end_row=end_row, end_column=5)
    iu_c.merge_cells(start_row=empty_row, start_column=6, end_row=end_row, end_column=6)
    iu_c.merge_cells(start_row=empty_row, start_column=7, end_row=end_row, end_column=7)
    iu_c.merge_cells(start_row=empty_row, start_column=8, end_row=end_row, end_column=8)
    iu_c.merge_cells(start_row=empty_row, start_column=9, end_row=end_row, end_column=9)
    iu_c.merge_cells(start_row=empty_row, start_column=10, end_row=end_row, end_column=10)
    iu_c.merge_cells(start_row=empty_row, start_column=13, end_row=end_row, end_column=13)
    iu_c.merge_cells(start_row=empty_row, start_column=14, end_row=end_row, end_column=14)
    iu_c.merge_cells(start_row=empty_row, start_column=15, end_row=end_row, end_column=15)
    iu_c.merge_cells(start_row=empty_row, start_column=16, end_row=end_row, end_column=16)
    iu_c.merge_cells(start_row=empty_row, start_column=17, end_row=end_row, end_column=17)
    iu_c.merge_cells(start_row=empty_row, start_column=18, end_row=end_row, end_column=18)
    iu_c.merge_cells(start_row=empty_row, start_column=19, end_row=end_row, end_column=19)
    iu_c.merge_cells(start_row=empty_row, start_column=20, end_row=end_row, end_column=20)
    iu_c.merge_cells(start_row=empty_row, start_column=21, end_row=end_row, end_column=21)
    iu_c.merge_cells(start_row=empty_row, start_column=22, end_row=end_row, end_column=22)
    iu_c.merge_cells(start_row=empty_row, start_column=29, end_row=end_row, end_column=29)
    iu_c.merge_cells(start_row=empty_row, start_column=30, end_row=end_row, end_column=30)
    iu_c.merge_cells(start_row=empty_row, start_column=31, end_row=end_row, end_column=31)
    iu_c.merge_cells(start_row=empty_row, start_column=32, end_row=end_row, end_column=32)
    iu_c.merge_cells(start_row=empty_row, start_column=41, end_row=end_row, end_column=41)
    iu_c.merge_cells(start_row=empty_row, start_column=42, end_row=end_row, end_column=42)
    iu_c.merge_cells(start_row=empty_row, start_column=43, end_row=end_row, end_column=43)
    iu_c['{}{}'.format(rnc_column_id.get('RNC Name'), empty_row)] = rnc_data[counter].get('RNCN')
    iu_c['{}{}'.format(rnc_column_id.get('Vendor'), empty_row)] = 'NA'
    iu_c['{}{}'.format(rnc_column_id.get('RNCX'), empty_row)] = rnc_data[counter].get('RNCX')
    iu_c['{}{}'.format(rnc_column_id.get('RNCID'), empty_row)] = rnc_data[counter].get('RNCID')
    iu_c['{}{}'.format(rnc_column_id.get('DPC'), empty_row)] = rnc_data[counter].get('DPC')
    iu_c['{}{}'.format(rnc_column_id.get('DPX'), empty_row)] = rnc_data[counter].get('DPX')
    iu_c['{}{}'.format(rnc_column_id.get('RNCMCC'), empty_row)] = rnc_data[counter].get('RNCMCC')
    iu_c['{}{}'.format(rnc_column_id.get('RNCMNC'), empty_row)] = rnc_data[counter].get('RNCMNC')
    iu_c['{}{}'.format(rnc_column_id.get('NI'), empty_row)] = 'NATB'
    iu_c['{}{}'.format(rnc_column_id.get('Iu-flex'), empty_row)] = rnc_data[counter].get('IU-FLEX')
    iu_c['{}{}'.format(rnc_column_id.get('RanSharing'), empty_row)] = 'YES'
    iu_c['{}{}'.format(rnc_column_id.get('Iu-flex'), empty_row)] = rnc_data[counter].get('IU-FLEX')
    iu_c['{}{}'.format(rnc_column_id.get('DT'), empty_row)] = 'YES'
    iu_c['{}{}'.format(rnc_column_id.get('Support R7 Qos'), empty_row)] = 'YES'
    iu_c['{}{}'.format(rnc_column_id.get('RAB QOS'), empty_row)] = rnc_data[counter].get('RABQOS')
    iu_c['{}{}'.format(rnc_column_id.get('RNC Version'), empty_row)] = 'R8'
    iu_c['{}{}'.format(rnc_column_id.get('SGSN Name'), empty_row)] = usn_data.get('USN')
    iu_c['{}{}'.format(rnc_column_id.get('NRI'), empty_row)] = usn_data.get('NRI')
    iu_c['{}{}'.format(rnc_column_id.get('OPC'), empty_row)] = usn_data.get('OPC')
    iu_c['{}{}'.format(rnc_column_id.get('OPX'), empty_row)] = '2'
    iu_c['{}{}'.format(rnc_column_id.get('NI Logical Link'), empty_row)] = 'NATB'
    iu_c['{}{}'.format(rnc_column_id.get('LEX'), empty_row)] = rnc_data[counter].get('LEX')
    iu_c['{}{}'.format(rnc_column_id.get('DEX'), empty_row)] = rnc_data[counter].get('DEX')
    iu_c['{}{}'.format(rnc_column_id.get('LSX'), empty_row)] = rnc_data[counter].get('LSX')
    iu_c['{}{}'.format(rnc_column_id.get('DET'), empty_row)] = rnc_data[counter].get('DET')
    iu_c['{}{}'.format(rnc_column_id.get('SSNX (RANAP)'), empty_row)] = rnc_data[counter].get('SSNX_RANAP')
    iu_c['{}{}'.format(rnc_column_id.get('SSNX (SCMG)'), empty_row)] = rnc_data[counter].get('SSNX_SCMG')
    iu_c['{}{}'.format(rnc_column_id.get('SLSM'), empty_row)] = len(rnc_data[counter].get('M3LNK'))
    max_lai = len(rnc_data[counter].get('LAI'))
    lai_count = 0
    for row_count in xrange(empty_row, empty_row + max_lai):
        iu_c['{}{}'.format(rnc_column_id.get('LAC'), row_count)] = rnc_data[counter].get('LAI')[lai_count].get(
            'LAC')
        iu_c['{}{}'.format(rnc_column_id.get('RAC'), row_count)] = rnc_data[counter].get('LAI')[lai_count].get(
            'RAC')
        lai_count += 1

    max_m3lnk = len(rnc_data[counter].get('M3LNK'))
    m3lnk_count = 0

    for row_count in xrange(empty_row, empty_row + max_m3lnk):
        esu_subrack.extend(list(roundrobin('0', '1')))
        usn_bayface = bayface.get(usn_data.get('USN'))
        rack_count = normalizando_row_count(row_count, max_m3lnk)
        tamanho = len(usn_bayface.get(esu_subrack[rack_count]))
        slot_chosen = choose_slot_number(usn_bayface,
                                         esu_subrack[rack_count],
                                         tamanho)
        iu_c['{}{}'.format(rnc_column_id.get('ESU Subrack'), row_count)] = esu_subrack[rack_count]
        iu_c['{}{}'.format(rnc_column_id.get('ESU Slot'), row_count)] = slot_chosen[0]
        if esu_subrack[rack_count] == '0':
            slot_count_0 = slot_chosen[1]
        else:
            slot_count_1 = slot_chosen[1]
        iu_c['{}{}'.format(rnc_column_id.get('LNK'), row_count)] = rnc_data[counter].get('RNCX')
        iu_c['{}{}'.format(rnc_column_id.get('PRI'), row_count)] = '0'
        iu_c['{}{}'.format(rnc_column_id.get('Entity Type'), row_count)] = rnc_data[counter].get('DET')
        iu_c['{}{}'.format(rnc_column_id.get('RNC Entity Type'), row_count)] = rnc_data[counter].get('DET')
        iu_c['{}{}'.format(rnc_column_id.get('SGSN IP1'), row_count)] = \
            choose_sgsn_ip(esu_subrack[rack_count], usn_data, 'Iu_C')[0]
        iu_c['{}{}'.format(rnc_column_id.get('SGSN IP2'), row_count)] = \
            choose_sgsn_ip(esu_subrack[rack_count], usn_data, 'Iu_C')[1]
        iu_c['{}{}'.format(rnc_column_id.get('SGSN Port'), row_count)] = str(port_number)
        iu_c['{}{}'.format(rnc_column_id.get('RNC IP1'), row_count)] = rnc_data[counter].get(
            'M3LNK')[m3lnk_count].get('PEERIPV41')
        iu_c['{}{}'.format(rnc_column_id.get('RNC IP2'), row_count)] = rnc_data[counter].get(
            'M3LNK')[m3lnk_count].get('PEERIPV42')
        iu_c['{}{}'.format(rnc_column_id.get('RNC Port'), row_count)] = str(port_number)
        iu_c['{}{}'.format(rnc_column_id.get('Share Type'), row_count)] = 'LOADSHARE'
        iu_c['{}{}'.format(rnc_column_id.get('C/S'), row_count)] = 'S'
        m3lnk_count += 1
        port_number += 1


def populate_gboip_sheet_with_bsc_info(gbo_ip, empty_row, end_row, bsc_data, counter, usn_data):
    # TODO Implement a better way to handle information from GBLOCALENDPOINT
    gbo_ip.merge_cells(start_row=empty_row, start_column=1, end_row=end_row, end_column=1)
    gbo_ip.merge_cells(start_row=empty_row, start_column=2, end_row=end_row, end_column=2)
    gbo_ip.merge_cells(start_row=empty_row, start_column=5, end_row=end_row, end_column=5)
    gbo_ip.merge_cells(start_row=empty_row, start_column=6, end_row=end_row, end_column=6)
    gbo_ip.merge_cells(start_row=empty_row, start_column=7, end_row=end_row, end_column=7)
    gbo_ip.merge_cells(start_row=empty_row, start_column=13, end_row=end_row, end_column=13)
    gbo_ip['{}{}'.format(bsc_column_id.get('BSC Name'), empty_row)] = bsc_data[counter].get('BSC Name')
    gbo_ip['{}{}'.format(bsc_column_id.get('TX'), empty_row)] = bsc_data[counter].get('BT')
    gbo_ip['{}{}'.format(bsc_column_id.get('NSE Conf Type'), empty_row)] = bsc_data[counter].get('CT')
    gbo_ip['{}{}'.format(bsc_column_id.get('SGSN Name'), empty_row)] = usn_data.get('USN')
    gbo_ip['{}{}'.format(bsc_column_id.get('NRI'), empty_row)] = usn_data.get('NRI')
    gbo_ip['{}{}'.format(bsc_column_id.get('GB-FLEX'), empty_row)] = bsc_data[counter].get('GB-FLEX')
    max_nsei = len(bsc_data[counter].get('NSE'))
    local_endpoint_count = 0
    nsei_count = 0
    for row_count in xrange(empty_row, empty_row + max_nsei):
        gbo_ip['{}{}'.format(bsc_column_id.get('NSEI'), row_count)] = bsc_data[counter].get('NSE')[nsei_count].get(
            'NSEI')
        gbo_ip['{}{}'.format(bsc_column_id.get('BSSID'), row_count)] = bsc_data[counter].get('NSE')[nsei_count].get(
            'BSSID')
        nsei_count += 1
    max_local_endpoint = len(bsc_data[counter].get('GBLOCALENDPOINT'))
    row_count_real = empty_row
    for row_count in xrange(empty_row, empty_row + max_local_endpoint):
        if not local_endpoint_count % 2:
            gbo_ip['{}{}'.format(bsc_column_id.get('Subrack No'), row_count_real)] = bsc_data[counter].get(
                'GBLOCALENDPOINT')[local_endpoint_count].get(
                'SRN')
            gbo_ip['{}{}'.format(bsc_column_id.get('Slot No'), row_count_real)] = bsc_data[counter].get(
                'GBLOCALENDPOINT')[local_endpoint_count].get(
                'SN')

            gbo_ip['{}{}'.format(bsc_column_id.get('Local IP Address 1'), row_count_real)] = bsc_data[counter].get(
                'GBLOCALENDPOINT')[local_endpoint_count].get(
                'LIPV4')
            gbo_ip['{}{}'.format(bsc_column_id.get('Local Port'), row_count_real)] = bsc_data[counter].get(
                'GBLOCALENDPOINT')[local_endpoint_count].get(
                'LUP')
            row_count_real += 1
        else:
            gbo_ip['{}{}'.format(bsc_column_id.get('Local IP Address 2'), row_count_real - 1)] = bsc_data[counter].get(
                'GBLOCALENDPOINT')[local_endpoint_count].get(
                'LIPV4')
        local_endpoint_count += 1



def populate_gboip_sheet_with_usn_info(gbo_ip, empty_row, end_row, bsc_data, counter, usn_data, esu_subrack,
                                       port_number, legacy_usn_data):
    global slot_count_0
    global slot_count_1
    gbo_ip.merge_cells(start_row=empty_row, start_column=1, end_row=end_row, end_column=1)
    gbo_ip.merge_cells(start_row=empty_row, start_column=2, end_row=end_row, end_column=2)
    gbo_ip.merge_cells(start_row=empty_row, start_column=5, end_row=end_row, end_column=5)
    gbo_ip.merge_cells(start_row=empty_row, start_column=6, end_row=end_row, end_column=6)
    gbo_ip.merge_cells(start_row=empty_row, start_column=7, end_row=end_row, end_column=7)
    gbo_ip.merge_cells(start_row=empty_row, start_column=13, end_row=end_row, end_column=13)
    gbo_ip['{}{}'.format(bsc_column_id.get('BSC Name'), empty_row)] = bsc_data[counter].get('BSC Name')
    gbo_ip['{}{}'.format(bsc_column_id.get('TX'), empty_row)] = bsc_data[counter].get('BT')
    gbo_ip['{}{}'.format(bsc_column_id.get('NSE Conf Type'), empty_row)] = bsc_data[counter].get('CT')
    gbo_ip['{}{}'.format(bsc_column_id.get('SGSN Name'), empty_row)] = usn_data.get('USN')
    gbo_ip['{}{}'.format(bsc_column_id.get('NRI'), empty_row)] = usn_data.get('NRI')
    gbo_ip['{}{}'.format(bsc_column_id.get('GB-FLEX'), empty_row)] = bsc_data[counter].get('GB-FLEX')
    max_nsei = len(bsc_data[counter].get('NSE'))
    nsei_count = 0
    row_count_real = empty_row
    for row_count in xrange(empty_row, empty_row + max_nsei):
        esu_subrack.extend(list(roundrobin('0', '1')))
        usn_bayface = bayface.get(usn_data.get('USN'))
        rack_count = normalizando_row_count(row_count, max_nsei)
        tamanho = len(usn_bayface.get(esu_subrack[rack_count]))
        slot_chosen = choose_slot_number(usn_bayface,
                                         esu_subrack[rack_count],
                                         tamanho)
        nsei = list(bsc_data[counter].get('NSE')[nsei_count].get('NSEI'))
        if nsei[0] == nsei_start_number.get(legacy_usn_data.get('USN'))[0]:
            nsei[0] = nsei_start_number.get(usn_data.get('USN'))[0]
        elif nsei[0] == nsei_start_number.get(legacy_usn_data.get('USN'))[1]:
            nsei[0] = nsei_start_number.get(usn_data.get('USN'))[1]
        elif int(nsei[0]) % 2:
            nsei[0] = nsei_start_number.get(usn_data.get('USN'))[1]
        else:
            nsei[0] = nsei_start_number.get(usn_data.get('USN'))[0]
        nsei_string = ''
        gbo_ip['{}{}'.format(bsc_column_id.get('NSEI'), row_count)] = nsei_string.join(nsei)
        gbo_ip['{}{}'.format(bsc_column_id.get('BSSID'), row_count)] = nsei_string.join(nsei)

        gbo_ip['{}{}'.format(bsc_column_id.get('Subrack No'), row_count_real)] = esu_subrack[rack_count]
        gbo_ip['{}{}'.format(bsc_column_id.get('Slot No'), row_count_real)] = slot_chosen[0]
        if esu_subrack[rack_count] == '0':
            slot_count_0 = slot_chosen[1]
        else:
            slot_count_1 = slot_chosen[1]

        gbo_ip['{}{}'.format(bsc_column_id.get('Local IP Address 1'), row_count_real)] = choose_sgsn_ip(
            esu_subrack[rack_count], usn_data, 'Gb')[0]
        gbo_ip['{}{}'.format(bsc_column_id.get('Local IP Address 2'), row_count_real)] = choose_sgsn_ip(
            esu_subrack[rack_count], usn_data, 'Gb')[1]
        gbo_ip['{}{}'.format(bsc_column_id.get('Local Port'), row_count_real)] = str(port_number)

        row_count_real += 1
        port_number += 5
        nsei_count += 1

    return port_number


def choose_slot_number(usn_bayface, subrack, tamanho):
    global slot_count_0
    global slot_count_1
    if subrack == '0':
        slot_count = slot_count_0 % tamanho
        slot_number = usn_bayface.get(subrack)[slot_count]
        slot_count_0 += 1
        return slot_number, slot_count_0
    elif subrack == '1':
        slot_count = slot_count_1 % tamanho
        slot_number = usn_bayface.get(subrack)[slot_count]
        slot_count_1 += 1
        return slot_number, slot_count_1


def choose_sgsn_ip(subrack, usn_data, usn_data_key):
    ip_primario = ''
    ip_secundario = ''
    if subrack == '0':
        lista_ips = usn_data.get(usn_data_key)[0:2]
        ip_primario = lista_ips[0].get('IP')
        ip_secundario = lista_ips[1].get('IP')
    elif subrack == '1':
        lista_ips = usn_data.get(usn_data_key)[2:]
        ip_primario = lista_ips[0].get('IP')
        ip_secundario = lista_ips[1].get('IP')

    return ip_primario, ip_secundario


def normalizando_row_count(row_count, max_row):
    if row_count % max_row == 0:
        rack_count = max_row - 1
    else:
        rack_count = (row_count % max_row) - 1

    return rack_count


if __name__ == '__main__':
    main()
